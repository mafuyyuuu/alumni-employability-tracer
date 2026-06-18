import os
import re
import hashlib
import random
import string
import smtplib
import threading
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from flask import Blueprint, jsonify, request, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from database import get_db
from ml.arima_model import run_arima_forecast, parse_order
from ml.train_lr import run_lr_forecast
from ml.rf_forecast import run_rf_forecast
from ml.predictor import (
    predict_employability_details,
    predictor_feature_importance,
    predictor_status,
    ml_predictor,
)
from ml.dataset_importer import (
    import_first_clean_dataset as import_first_clean_dataset_rows,
    import_training_csv,
    COURSE_ALIASES,
)
from ml.train_rf import train_random_forest
from ml.train_employability_lr import train_linear_employability
from ml.training_data import (
    load_training_dataframe,
    validate_training_dataframe,
    build_feature_matrix,
)
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LinearRegression
from functools import wraps

admin_bp = Blueprint('admin', __name__)

# ── Background job state ──────────────────────────────────────────────────────
_training_job = {'status': 'idle', 'result': None, 'error': None}
_training_lock = threading.Lock()

_upload_progress = {'stage': 'idle', 'percent': 0, 'message': '', 'total': 0, 'done': 0}
_upload_progress_lock = threading.Lock()

def _set_progress(stage, percent, message='', total=0, done=0):
    with _upload_progress_lock:
        _upload_progress.update({'stage': stage, 'percent': percent,
                                  'message': message, 'total': total, 'done': done})

def _run_training_background(app, db_path, dataset_year):
    global _training_job
    try:
        with app.app_context():
            _set_progress('training', 65, 'Training Random Forest…')
            rf_metadata = train_random_forest(database_path=db_path)
            _set_progress('training', 82, 'Training Linear Regression…')
            lr_metadata = train_linear_employability(database_path=db_path)
            _set_progress('training', 93, 'Loading models…')
            ml_predictor._load_models()
            forecast = None
            if dataset_year:
                try:
                    _set_progress('training', 96, 'Computing forecast…')
                    forecast = _auto_forecast_3yr(get_db(), dataset_year)
                except Exception:
                    forecast = None
            _set_progress('done', 100, 'Upload complete')
            with _training_lock:
                _training_job = {
                    'status': 'done',
                    'result': {'rf': rf_metadata, 'lr': lr_metadata, 'forecast': forecast},
                    'error': None,
                }
    except Exception as exc:
        _set_progress('error', 0, str(exc))
        with _training_lock:
            _training_job = {'status': 'error', 'result': None, 'error': str(exc)}


def _run_import_and_training_background(app, db_path, file_path, safe_name,
                                         dataset_year, retrain, upload_id,
                                         create_accounts, skip_email,
                                         conflict_mode, overwrite_all=False):
    """Run CSV/Excel import + optional retraining fully in background."""
    global _training_job
    try:
        with app.app_context():
            db = get_db()

            if overwrite_all:
                db.execute("DELETE FROM ml_training_rows")
                db.commit()
            elif conflict_mode == 'overwrite' and dataset_year:
                db.execute("DELETE FROM ml_training_rows WHERE graduation_year = ?", [dataset_year])
                db.commit()

            def _prog(done, total):
                pct = 5 + int((done / total) * 55)
                _set_progress('importing', pct, f'Importing rows… {done}/{total}', total, done)

            _set_progress('importing', 5, 'Importing data…')
            import_result = import_training_csv(
                database_path=db_path,
                csv_path=file_path,
                source_name=safe_name,
                year_override=dataset_year,
                progress_callback=_prog,
            )
            _set_progress('importing', 60, 'Import complete')
            db.execute(
                "UPDATE model_uploads SET applied_to_training = 1, status = 'Imported' WHERE id = ?",
                [upload_id],
            )
            db.commit()

            # Create alumni accounts if requested
            if create_accounts and dataset_year:
                try:
                    _create_alumni_accounts_from_dataset(
                        db, dataset_year, skip_email=skip_email
                    )
                except Exception:
                    pass

            if retrain:
                import time as _time
                _set_progress('training', 62, 'Starting model training…')
                with _training_lock:
                    _training_job = {'status': 'running', 'result': None, 'error': None}

                # RF training — emit ticks so frontend doesn't appear stuck at 62%
                _set_progress('training', 64, 'Training Random Forest (this may take a minute)…')
                _time.sleep(0.5)
                _set_progress('training', 66, 'Training Random Forest…')
                rf_metadata = train_random_forest(database_path=db_path)
                _set_progress('training', 80, 'Random Forest done. Training Linear Regression…')

                # LR training
                _time.sleep(0.3)
                _set_progress('training', 82, 'Training Linear Regression…')
                lr_metadata = train_linear_employability(database_path=db_path)
                _set_progress('training', 92, 'Linear Regression done. Loading models…')

                ml_predictor._load_models()
                _set_progress('training', 95, 'Models loaded. Computing forecast…')

                forecast = None
                if dataset_year:
                    try:
                        forecast = _auto_forecast_3yr(get_db(), dataset_year)
                    except Exception:
                        forecast = None

                _set_progress('done', 100, 'Upload and training complete!')
                with _training_lock:
                    _training_job = {
                        'status': 'done',
                        'result': {'rf': rf_metadata, 'lr': lr_metadata,
                                   'forecast': forecast, 'import': import_result},
                        'error': None,
                    }
            else:
                _set_progress('done', 100, 'Import complete!')
                with _training_lock:
                    _training_job = {
                        'status': 'done',
                        'result': {'import': import_result},
                        'error': None,
                    }
    except Exception as exc:
        _set_progress('error', 0, str(exc))
        with _training_lock:
            _training_job = {'status': 'error', 'result': None, 'error': str(exc)}


def _send_welcome_email(to_email, name, password):
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', 587))
    smtp_user = os.environ.get('SMTP_USER', '')
    smtp_pass = os.environ.get('SMTP_PASS', '')
    smtp_from = os.environ.get('SMTP_FROM', smtp_user)
    if not smtp_user or not smtp_pass:
        return False, 'SMTP not configured'
    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'Your PLP Alumni Portal Account'
    msg['From'] = smtp_from
    msg['To'] = to_email
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:24px;border:1px solid #e5e7eb;border-radius:12px">
      <div style="background:#163d22;border-radius:8px;padding:18px 24px;margin-bottom:20px">
        <h2 style="color:#fff;margin:0;font-size:18px">PLP Alumni Employability Portal</h2>
        <p style="color:#b7e4c7;margin:4px 0 0;font-size:13px">Pamantasan ng Lungsod ng Pasig</p>
      </div>
      <p style="color:#374151;font-size:14px">Hello <strong>{name}</strong>,</p>
      <p style="color:#374151;font-size:14px">Your alumni account has been created. Use the credentials below to log in:</p>
      <div style="background:#f3f4f6;border-radius:8px;padding:16px;margin:16px 0">
        <p style="margin:0 0 8px;font-size:13px;color:#6b7280">Email</p>
        <p style="margin:0 0 16px;font-size:15px;font-weight:700;color:#111827">{to_email}</p>
        <p style="margin:0 0 8px;font-size:13px;color:#6b7280">Temporary Password</p>
        <p style="margin:0;font-size:18px;font-weight:900;color:#163d22;letter-spacing:2px">{password}</p>
      </div>
      <p style="color:#6b7280;font-size:12px">Please change your password after your first login. If you did not expect this email, please contact your administrator.</p>
    </div>
    """
    msg.attach(MIMEText(html, 'html'))
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_from, to_email, msg.as_string())
        return True, None
    except Exception as e:
        return False, str(e)


def admin_required(fn):
    @wraps(fn)
    @jwt_required()
    def wrapper(*args, **kwargs):
        uid = get_jwt_identity()
        db = get_db()
        user = db.execute('SELECT role FROM users WHERE id = ?', [uid]).fetchone()
        if not user or user['role'] != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return fn(*args, **kwargs)
    return wrapper


def _alumni_features_from_db(db, user_id):
    row = db.execute("""
        SELECT
            course,
            graduation_year,
            age,
            avg_grade,
            avg_prof_grade,
            avg_elec_grade,
            ojt_grade,
            soft_skills,
            hard_skills,
            board_passer,
            board_exam_score
        FROM users
        WHERE id = ? AND role = 'alumni'
    """, [user_id]).fetchone()
    if not row:
        return None
    keys = row.keys()
    return {
        'course': row['course'],
        'graduation_year': row['graduation_year'],
        'age': row['age'],
        'avg_grade': row['avg_grade'],
        'avg_prof_grade': row['avg_prof_grade'],
        'avg_elec_grade': row['avg_elec_grade'],
        'ojt_grade': row['ojt_grade'],
        'soft_skills': row['soft_skills'],
        'hard_skills': row['hard_skills'],
        'board_passer': int(row['board_passer']) if 'board_passer' in keys else 0,
        'board_exam_score': float(row['board_exam_score']) if 'board_exam_score' in keys else 0.0,
    }


def _file_sha256(path):
    digest = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            digest.update(chunk)
    return digest.hexdigest()


def _clamp(value, lo=0.0, hi=1.0):
    return max(lo, min(hi, value))


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _normalize_course_value(value):
    raw = str(value or '').strip()
    if not raw:
        return ''
    upper = raw.upper()
    return COURSE_ALIASES.get(upper, raw)


def _parse_employment_flag(value):
    text = str(value or '').strip().lower()
    if text in ('1', 'true', 'yes', 'y', 'employed', 'hired', 'working'):
        return 1
    if text in ('0', 'false', 'no', 'n', 'unemployed', 'looking', 'seeking'):
        return 0
    try:
        numeric = int(float(text))
        if numeric in (0, 1):
            return numeric
    except (TypeError, ValueError):
        pass
    return None


def _is_truthy(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in ('1', 'true', 'yes', 'on'):
        return True
    if text in ('0', 'false', 'no', 'off'):
        return False
    return default


def _feature_number(features, *keys, default=0.0):
    for key in keys:
        if key in features and features[key] is not None:
            return _to_float(features[key], default)
    return float(default)


def _format_percent_metric(value):
    return f"{value}%" if isinstance(value, (int, float)) else str(value)


def _refresh_program_rates(db, year):
    if year is None:
        return []
    rows = db.execute("""
        SELECT course, COUNT(*) AS total, COALESCE(SUM(employed), 0) AS employed_count
        FROM ml_training_rows
        WHERE is_active = 1 AND graduation_year = ? AND course IS NOT NULL AND course != ''
        GROUP BY course
        ORDER BY course
    """, [year]).fetchall()
    db.execute("DELETE FROM program_rates WHERE year = ?", [year])
    for row in rows:
        total = row['total'] or 0
        if total <= 0:
            continue
        rate = round((row['employed_count'] / total) * 100, 2)
        db.execute(
            "INSERT INTO program_rates (year, course, rate) VALUES (?,?,?)",
            [year, row['course'], rate]
        )
    db.commit()
    return rows


def _get_program_rates(db, year):
    rows = db.execute(
        "SELECT course, rate FROM program_rates WHERE year = ? ORDER BY rate DESC",
        [year]
    ).fetchall()
    if rows:
        return rows
    _refresh_program_rates(db, year)
    return db.execute(
        "SELECT course, rate FROM program_rates WHERE year = ? ORDER BY rate DESC",
        [year]
    ).fetchall()


def _ensure_employment_data_from_training(db):
    rows = db.execute("""
        SELECT graduation_year AS year,
               COUNT(*) AS total,
               COALESCE(SUM(employed), 0) AS employed_count
        FROM ml_training_rows
        WHERE is_active = 1
        GROUP BY graduation_year
        ORDER BY graduation_year
    """).fetchall()
    if not rows:
        return False

    for row in rows:
        total = row['total'] or 0
        if total <= 0:
            continue
        employed_count = int(row['employed_count'])
        rate = round((employed_count / total) * 100, 2)
        db.execute("""
            INSERT INTO employment_data (year, overall_rate, employed_count, unemployed_count)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(year) DO UPDATE SET
                overall_rate = excluded.overall_rate,
                employed_count = excluded.employed_count,
                unemployed_count = excluded.unemployed_count
        """, [row['year'], rate, employed_count, int(total - employed_count)])
    db.commit()
    return True


def _format_factor_label(feature):
    if not feature:
        return 'Unknown'
    text = str(feature)
    
    degree_mapping = {
        'BSCS': 'Bachelor of Science in Computer Science',
        'BSIT': 'Bachelor of Science in Information Technology',
        'BSCPE': 'Bachelor of Science in Computer Engineering',
        'BSECE': 'Bachelor of Science in Electronics Engineering',
        'BSCE': 'Bachelor of Science in Civil Engineering',
        'BSN': 'Bachelor of Science in Nursing',
        'BSED': 'Bachelor of Secondary Education',
        'BEED': 'Bachelor of Elementary Education',
        'BSA': 'Bachelor of Science in Accountancy',
        'BSBA': 'Bachelor of Science in Business Administration',
        'BSHM': 'Bachelor of Science in Hospitality Management',
        'BSENTREP': 'Bachelor of Science in Entrepreneurship',
        'PSYCHOLOGY': 'Bachelor of Arts in Psychology',
        'BSIS': 'Bachelor of Science in Information Systems',
    }

    if text.startswith('course_'):
        code = text.replace('course_', '').upper()
        return f"Degree: {degree_mapping.get(code, code)}"

    mapping = {
        'avg_grade':      'General Weighted Average (GWA)',
        'avg_prof_grade': 'Professional Subject Performance',
        'avg_elec_grade': 'Major/Elective Excellence',
        'ojt_grade':      'On-the-Job Training (OJT) Rating',
        'soft_skills':    'Behavioral & Soft Skills',
        'hard_skills':    'Technical & Industry Skills',
        'board_passer':   'Licensure / Board Exam Status',
        'board_exam_score': 'Board Examination Rating',
        'graduation_year': 'Year of Graduation',
    }
    if text in mapping:
        return mapping[text]
    return text.replace('_', ' ').title()


def _importance_from_loaded_model(model_key):
    model_key = (model_key or 'rf').strip().lower()
    if model_key not in ('rf', 'lr'):
        return None, "Invalid model key. Use 'rf' or 'lr'."

    model_obj = ml_predictor.models.get(model_key)
    feature_names = ml_predictor.features.get(model_key)
    if model_obj is None or not feature_names:
        return None, f"Model '{model_key}' is not loaded."

    if model_key == 'rf':
        if not hasattr(model_obj, 'feature_importances_'):
            return None, "Random Forest model does not expose feature importances."
        raw = model_obj.feature_importances_
    else:
        if not hasattr(model_obj, 'coef_'):
            return None, "Linear Regression model does not expose coefficients."
        raw = model_obj.coef_

    mapping = {
        feature: abs(float(raw[idx]))
        for idx, feature in enumerate(feature_names)
        if idx < len(raw)
    }
    return mapping, None


def _importance_from_program_data(model_key, program_code):
    model_key = (model_key or 'rf').strip().lower()
    if model_key not in ('rf', 'lr'):
        return None, "Invalid model key. Use 'rf' or 'lr'."

    df = load_training_dataframe()
    if df.empty:
        return None, 'No training data available.'

    program = str(program_code or '').strip().upper()
    if not program:
        return None, 'Program code is required.'

    df = df[df['course'] == program]
    if df.empty:
        return None, f'No training rows available for {program}.'

    validate_training_dataframe(df, min_rows=10)
    X, y, _ = build_feature_matrix(df)
    feature_cols = [c for c in X.columns if not c.startswith('course_')]
    if not feature_cols:
        return None, f'No usable feature columns available for {program}.'

    X = X[feature_cols]
    stratify = y if y.value_counts().min() >= 2 else None
    if model_key == 'rf':
        model_obj = RandomForestClassifier(
            n_estimators=200,
            random_state=42,
            class_weight='balanced',
        )
        model_obj.fit(X, y)
        raw = model_obj.feature_importances_
    else:
        model_obj = LinearRegression()
        model_obj.fit(X, y)
        raw = model_obj.coef_

    mapping = {
        feature: abs(float(raw[idx]))
        for idx, feature in enumerate(feature_cols)
        if idx < len(raw)
    }
    return mapping, None


def _normalize_grade(value):
    return _clamp(_to_float(value) / 100.0)


def _normalize_age(value):
    age = _to_float(value, 22.0)
    # Expected alumni age range for scaling.
    return _clamp((age - 18.0) / 27.0)


def _normalize_model_choice(raw):
    text = (raw or '').strip().lower()
    if text in ('rf', 'random forest', 'random forest regressor', 'random forest classifier'):
        return 'rf'
    if text in ('lr', 'linear regression', 'linear'):
        return 'lr'
    if text.startswith('arima') or 'auto arima' in text:
        return 'arima'
    return 'rf'


def _forecast_result_for_model(rates, horizon, model_str):
    normalized = _normalize_model_choice(model_str)
    if normalized == 'rf':
        return run_rf_forecast(rates, horizon=horizon)
    if normalized == 'lr':
        return run_lr_forecast(rates, horizon=horizon)
    order = parse_order(model_str)
    return run_arima_forecast(rates, horizon=horizon, order=order)


def _predict_with_arima_employability(db, features):
    _ensure_employment_data_from_training(db)
    row = db.execute(
        "SELECT year, overall_rate FROM employment_data ORDER BY year"
    ).fetchall()
    if not row:
        return {'error': 'Employment trend data is unavailable for ARIMA prediction.'}

    rates = [float(r['overall_rate']) for r in row]
    years = [int(r['year']) for r in row]
    grad_year = int(_feature_number(features, 'graduation_year', 'graduationYear', default=years[-1]))

    if grad_year in years:
        rate = rates[years.index(grad_year)]
    elif grad_year < years[0]:
        rate = rates[0]
    else:
        horizon = max(grad_year - years[-1], 1)
        forecast = run_arima_forecast(rates, horizon=horizon, order=None)
        rate = float(forecast['forecast_values'][-1])

    probability = _clamp(rate / 100.0)
    prediction = 1 if probability >= 0.5 else 0
    return {
        'label': 'Employed' if prediction == 1 else 'Unemployed',
        'prediction': prediction,
        'probability_employed': round(probability, 4),
        'model_used': 'arima',
        'requested_model': 'arima',
        'model_note': 'ARIMA employability mode estimates cohort-level probability from historical employment trend.',
    }


def _read_prediction_settings(db):
    row = db.execute(
        "SELECT use_voter_weights FROM prediction_settings WHERE id = 1"
    ).fetchone()
    if not row:
        db.execute(
            "INSERT INTO prediction_settings (id, use_voter_weights) VALUES (1, 0)"
        )
        db.commit()
        return {'use_voter_weights': False}
    return {'use_voter_weights': bool(row['use_voter_weights'])}


def _read_voter_fields(db):
    rows = db.execute(
        "SELECT * FROM voter_config ORDER BY id"
    ).fetchall()
    return [{
        'id': r['id'],
        'name': r['field_name'],
        'key': r['field_key'],
        'enabled': bool(r['enabled']),
        'weight': int(r['weight'] or 0),
    } for r in rows]


def _normalize_weights_to_100(raw_weights):
    keys = list(raw_weights.keys())
    total = sum(max(0.0, float(v)) for v in raw_weights.values())
    if total <= 0:
        return {key: 0 for key in keys}

    scaled = {key: (max(0.0, float(raw_weights[key])) / total) * 100.0 for key in keys}
    rounded = {key: int(scaled[key]) for key in keys}
    remainder = 100 - sum(rounded.values())

    if remainder != 0:
        fractions = sorted(
            ((key, scaled[key] - int(scaled[key])) for key in keys),
            key=lambda item: item[1],
            reverse=(remainder > 0),
        )
        idx = 0
        while remainder != 0 and fractions:
            key = fractions[idx % len(fractions)][0]
            if remainder > 0:
                rounded[key] += 1
                remainder -= 1
            else:
                if rounded[key] > 0:
                    rounded[key] -= 1
                    remainder += 1
            idx += 1

    return rounded


def _suggest_voter_weights_from_rf(voter_fields):
    importance_result = predictor_feature_importance(model='rf')
    if importance_result.get('error'):
        return {'error': importance_result['error']}

    feature_importance = importance_result.get('feature_importance', {})
    factor_feature_map = {
        'gpa': ('avg_grade',),
        'prof_grade': ('avg_prof_grade',),
        'elec_grade': ('avg_elec_grade',),
        'ojt_grade': ('ojt_grade',),
        'soft_skills': ('soft_skills',),
        'hard_skills': ('hard_skills',),
        'age': ('age',),
        # Keep gender neutral/unmapped for ML-suggested weights.
        'gender': (),
    }

    raw_factor_weights = {}
    for field in voter_fields:
        if not field.get('enabled'):
            continue
        key = field['key']
        mapped_features = factor_feature_map.get(key, ())
        raw_factor_weights[key] = sum(feature_importance.get(feature, 0.0) for feature in mapped_features)

    if not raw_factor_weights:
        return {'error': 'No enabled voter factors available for ML suggestion.'}

    normalized = _normalize_weights_to_100(raw_factor_weights)
    suggested = []
    for field in voter_fields:
        key = field['key']
        if field.get('enabled'):
            suggested_weight = normalized.get(key, 0)
        else:
            suggested_weight = int(field.get('weight', 0) or 0)

        suggested.append({
            **field,
            'weight': max(0, min(100, int(suggested_weight))),
        })

    return {
        'config': suggested,
        'model': 'rf',
        'top_features': importance_result.get('top_features', []),
    }


def _predict_with_voter_weights(features, voter_fields):
    enabled = [f for f in voter_fields if f.get('enabled') and int(f.get('weight', 0)) > 0]
    if not enabled:
        return {
            'label': 'Employed',
            'prediction': 1,
            'probability_employed': 0.5,
            'reason': 'No enabled voter factors with positive weights.',
        }

    factor_scores = {}
    for field in enabled:
        key = field['key']
        if key == 'gpa':
            factor_scores[key] = _normalize_grade(_feature_number(features, 'avg_grade', 'avgGrade'))
        elif key == 'prof_grade':
            factor_scores[key] = _normalize_grade(_feature_number(features, 'avg_prof_grade', 'avgProfGrade'))
        elif key == 'elec_grade':
            factor_scores[key] = _normalize_grade(_feature_number(features, 'avg_elec_grade', 'avgElecGrade'))
        elif key == 'ojt_grade':
            factor_scores[key] = _normalize_grade(_feature_number(features, 'ojt_grade', 'ojtGrade'))
        elif key == 'soft_skills':
            factor_scores[key] = _normalize_grade(_feature_number(features, 'soft_skills', 'softSkills'))
        elif key == 'hard_skills':
            factor_scores[key] = _normalize_grade(_feature_number(features, 'hard_skills', 'hardSkills'))
        elif key == 'age':
            factor_scores[key] = _normalize_age(_feature_number(features, 'age'))
        elif key == 'board_passer':
            raw = _feature_number(features, 'board_passer', 'boardPasser')
            factor_scores[key] = 1.0 if raw >= 1 else 0.0
        elif key == 'gender':
            factor_scores[key] = 0.5
        else:
            factor_scores[key] = 0.5

    total_weight = sum(int(f['weight']) for f in enabled)
    weighted_sum = sum(factor_scores[f['key']] * int(f['weight']) for f in enabled)
    probability = _clamp(weighted_sum / total_weight if total_weight > 0 else 0.5)
    prediction = 1 if probability >= 0.5 else 0
    return {
        'label': 'Employed' if prediction == 1 else 'Unemployed',
        'prediction': prediction,
        'probability_employed': round(probability, 4),
        'reason': 'Computed from configured voter factor weights.',
    }


@admin_bp.route('/data-health', methods=['GET'])
@admin_required
def data_health():
    df = load_training_dataframe()
    if df.empty:
        return jsonify({'error': 'No training data available.'}), 200

    total_rows = len(df)
    health_metrics = []
    
    critical_fields = ['age', 'avg_grade', 'soft_skills', 'hard_skills', 'ojt_grade', 'course']
    
    for field in critical_fields:
        if field not in df.columns:
            continue
        missing = df[field].isna().sum()
        if field == 'course':
            missing += (df[field] == 'UNKNOWN').sum()
        
        missing_pct = round((missing / total_rows) * 100, 1)
        
        # Outlier detection (simple IQR)
        outliers = 0
        if field in NUMERIC_FEATURES and not df[field].empty:
            Q1 = df[field].quantile(0.25)
            Q3 = df[field].quantile(0.75)
            IQR = Q3 - Q1
            outliers = ((df[field] < (Q1 - 1.5 * IQR)) | (df[field] > (Q3 + 1.5 * IQR))).sum()

        health_metrics.append({
            'field': field.replace('_', ' ').title(),
            'missing_pct': missing_pct,
            'outliers': int(outliers),
            'status': 'Good' if missing_pct < 5 else 'Warning' if missing_pct < 15 else 'Critical'
        })

    # Class balance
    employed_pct = round((df['employed'] == 1).sum() / total_rows * 100, 1)
    unemployed_pct = 100 - employed_pct

    return jsonify({
        'total_rows': total_rows,
        'metrics': health_metrics,
        'balance': {
            'employed_pct': employed_pct,
            'unemployed_pct': unemployed_pct,
            'status': 'Balanced' if 30 <= employed_pct <= 70 else 'Imbalanced'
        }
    }), 200


# ── Dashboard ──────────────────────────────────────────────────────────────

@admin_bp.route('/dashboard', methods=['GET'])
@admin_required
def dashboard():
    db = get_db()
    # Auto-pick best model by R² — no manual selection on dashboard
    model_str = 'Linear Regression'

    # Check if any real dataset has been uploaded
    training_count = db.execute(
        "SELECT COUNT(*) as cnt FROM ml_training_rows WHERE is_active=1"
    ).fetchone()['cnt']

    has_dataset = training_count > 0

    # Total alumni = from dataset when available, else registered users (exclude test accounts)
    total_alumni = training_count if has_dataset else db.execute(
        "SELECT COUNT(*) as cnt FROM users WHERE role = 'alumni' AND (is_test_account = 0 OR is_test_account IS NULL)"
    ).fetchone()['cnt']

    # Stats are only meaningful when a real dataset exists
    employment_rate = 0
    graduate_success = 0
    emp_change = 0.0
    margin_of_error = 0
    chart_data = []

    if has_dataset:
        latest_yr = db.execute(
            "SELECT MAX(graduation_year) AS my FROM ml_training_rows WHERE is_active=1"
        ).fetchone()['my']

        employed_count = db.execute(
            "SELECT COUNT(*) as cnt FROM ml_training_rows WHERE is_active=1 AND employed=1"
            + (" AND graduation_year != ?" if latest_yr else ""),
            [latest_yr] if latest_yr else []
        ).fetchone()['cnt']

        non_graduating = db.execute(
            "SELECT COUNT(*) as cnt FROM ml_training_rows WHERE is_active=1"
            + (" AND graduation_year != ?" if latest_yr else ""),
            [latest_yr] if latest_yr else []
        ).fetchone()['cnt']

        # Use the most recent year's rate from employment_data so card matches graph
        _ensure_employment_data_from_training(db)
        latest_emp_row = db.execute(
            "SELECT overall_rate FROM employment_data ORDER BY year DESC LIMIT 1"
        ).fetchone()
        employment_rate = round(float(latest_emp_row['overall_rate']), 1) if latest_emp_row else (
            round(employed_count / non_graduating * 100, 1) if non_graduating > 0 else 0
        )

        # Graduate success = employment rate of the latest uploaded year's cohort
        latest_yr_total = db.execute(
            "SELECT COUNT(*) as cnt FROM ml_training_rows WHERE is_active=1 AND graduation_year=?",
            [latest_yr]
        ).fetchone()['cnt'] if latest_yr else 0
        latest_yr_employed = db.execute(
            "SELECT COUNT(*) as cnt FROM ml_training_rows WHERE is_active=1 AND employed=1 AND graduation_year=?",
            [latest_yr]
        ).fetchone()['cnt'] if latest_yr else 0
        graduate_success = round(latest_yr_employed / latest_yr_total * 100, 1) if latest_yr_total > 0 else 0

        _ensure_employment_data_from_training(db)
        emp_rows = db.execute(
            "SELECT year, overall_rate FROM employment_data ORDER BY year"
        ).fetchall()

        chart_data = [{'year': str(r['year']), 'rate': r['overall_rate']} for r in emp_rows]

        forecast = None
        if chart_data:
            rates = [r['overall_rate'] for r in emp_rows]
            forecast = _forecast_result_for_model(rates, horizon=1, model_str=model_str)
            next_year = str(emp_rows[-1]['year'] + 1)
            chart_data.append({
                'year': next_year,
                'rate': forecast['forecast_values'][0],
                'forecast': True,
            })
            mape = forecast.get('metrics', {}).get('mape')
            if isinstance(mape, (int, float)):
                margin_of_error = round(float(mape), 1)

        if len(emp_rows) >= 2:
            emp_change = round(float(emp_rows[-1]['overall_rate']) - float(emp_rows[-2]['overall_rate']), 1)

    return jsonify({
        'metrics': {
            'total_alumni': total_alumni,
            'employment_rate': employment_rate,
            'employment_rate_change': emp_change,
            'graduate_success': graduate_success,
            'margin_of_error': margin_of_error,
        },
        'employment_data': chart_data,
        'model_used': 'Linear Regression',
    }), 200


# ── Users ──────────────────────────────────────────────────────────────────

@admin_bp.route('/users', methods=['GET'])
@admin_required
def list_users():
    db = get_db()
    search       = request.args.get('search', '').lower()
    filter_by    = request.args.get('filter', 'All')
    year_filter  = request.args.get('year_filter', '')
    course_filter = request.args.get('course_filter', '')

    # Available years and programs for dropdown options
    available_years   = [r['graduation_year'] for r in db.execute(
        "SELECT DISTINCT graduation_year FROM ml_training_rows WHERE is_active=1 ORDER BY graduation_year DESC"
    ).fetchall()]
    available_courses = [r['course'] for r in db.execute(
        "SELECT DISTINCT course FROM ml_training_rows WHERE is_active=1 AND course IS NOT NULL ORDER BY course"
    ).fetchall()]

    # 1. Fetch registered alumni — exclude test accounts by flag OR by known test emails
    TEST_EMAILS = (
        'juan.delacruz@plp.edu.ph','maria.santos@plp.edu.ph','pedro.reyes@plp.edu.ph',
        'ana.garcia@plp.edu.ph','jose.mendoza@plp.edu.ph',
        'demo.alumni@plp.edu.ph','demo2.alumni@plp.edu.ph',
        'bscs.1@plp.edu.ph','bscs.2@plp.edu.ph',
        'bsit.1@plp.edu.ph','bsit.2@plp.edu.ph',
        'bscpe.1@plp.edu.ph','bscpe.2@plp.edu.ph',
        'bsece.1@plp.edu.ph','bsece.2@plp.edu.ph',
        'bsce.1@plp.edu.ph','bsce.2@plp.edu.ph',
        'bsn.1@plp.edu.ph','bsn.2@plp.edu.ph',
        'bsed.1@plp.edu.ph','bsed.2@plp.edu.ph',
        'beed.1@plp.edu.ph','beed.2@plp.edu.ph',
        'bsa.1@plp.edu.ph','bsa.2@plp.edu.ph',
        'bsba.1@plp.edu.ph','bsba.2@plp.edu.ph',
        'bshm.1@plp.edu.ph','bshm.2@plp.edu.ph',
    )
    placeholders = ','.join('?' * len(TEST_EMAILS))
    reg_rows = db.execute(
        f"SELECT * FROM users WHERE role='alumni' "
        f"AND (is_test_account=0 OR is_test_account IS NULL) "
        f"AND LOWER(email) NOT IN ({placeholders})",
        list(TEST_EMAILS)
    ).fetchall()
    reg_emails = {r['email'].lower() for r in reg_rows if r['email']}

    # 2. Fetch dataset rows not already registered
    ds_rows = db.execute("SELECT * FROM ml_training_rows WHERE is_active = 1").fetchall()

    all_rows = []
    
    # Add registered ones first
    for r in reg_rows:
        row_dict = dict(r)
        row_dict['type'] = 'Registered'
        all_rows.append(row_dict)

    # Add dataset ones if they don't overlap by email
    for r in ds_rows:
        email = (r['email'] or '').lower()
        if email and email in reg_emails:
            continue
        
        # Map dataset row to user-like structure for the UI
        ds_user = {
            'id': f"ds_{r['id']}",
            'type': 'Dataset',
            'first_name': r['name'] or f"Alumni {r['source_row_id']}",
            'last_name': '',
            'name': r['name'] or f"Alumni {r['source_row_id']}",
            'email': r['email'] or f"ID: {r['source_row_id']}",
            'course': r['course'],
            'graduation_year': r['graduation_year'],
            'age': r['age'],
            'account_status': 'Active',
            'employed': r['employed'],
            'avg_grade': r['avg_grade'],
            'soft_skills': r['soft_skills'],
            'hard_skills': r['hard_skills'],
            'ojt_grade': r['ojt_grade'],
            'avg_prof_grade': r['avg_prof_grade'],
            'avg_elec_grade': r['avg_elec_grade'],
            'board_passer': r['board_passer'],
            'board_exam_score': r['board_exam_score'],
            'months_to_employment': r['months_to_employment'],
            'ncae_completed': True,
        }
        all_rows.append(ds_user)

    SCORE_THRESHOLD = 75  # high score ≥ 75
    FAST_MONTHS     = 6   # fast employment ≤ 6 months

    # Latest uploaded dataset year = graduating students reference year
    lr = db.execute(
        "SELECT MAX(graduation_year) AS my FROM ml_training_rows WHERE is_active=1"
    ).fetchone()
    LATEST_YEAR = lr['my'] if (lr and lr['my']) else None

    # ── Historical k-NN data (employed alumni from training set with known months) ──
    hist_rows = db.execute("""
        SELECT course, avg_grade, soft_skills, hard_skills,
               ojt_grade, avg_prof_grade, avg_elec_grade, months_to_employment
        FROM ml_training_rows
        WHERE is_active=1 AND employed=1 AND months_to_employment IS NOT NULL
    """).fetchall()

    # Group by course for faster lookup
    _hist_by_course = {}
    for h in hist_rows:
        c = (h['course'] or '').upper().strip()
        _hist_by_course.setdefault(c, []).append(h)

    def _norm_gwa(g):
        g = float(g or 0)
        return round((5.0 - g) / 4.0 * 100, 1) if 0 < g <= 5.0 else g

    def _feats(r):
        return [
            _norm_gwa(r.get('avg_grade', 0)),
            float(r.get('soft_skills')    or 0),
            float(r.get('hard_skills')    or 0),
            float(r.get('ojt_grade')      or 0),
            float(r.get('avg_prof_grade') or 0),
            float(r.get('avg_elec_grade') or 0),
        ]

    def _knn_months(r, k=10):
        """Predict months-to-employment via k-NN against historical employed alumni."""
        course = (r.get('course') or '').upper().strip()
        pool   = _hist_by_course.get(course) or list(hist_rows)
        if len(pool) < k:
            pool = list(hist_rows)

        uf = _feats(r)
        dists = sorted(
            ((sum((a - float(h[col] or 0)) ** 2
                  for a, col in zip(uf, ['avg_grade','soft_skills','hard_skills',
                                         'ojt_grade','avg_prof_grade','avg_elec_grade'])) ** 0.5,
              int(h['months_to_employment']))
             for h in pool),
            key=lambda x: x[0]
        )[:k]

        if not dists:
            return None
        total_w  = sum(1.0 / (d + 1e-6) for d, _ in dists)
        predicted = sum(m / (d + 1e-6) for d, m in dists) / total_w
        return round(predicted)

    def _employability_score(r):
        avg_g = _norm_gwa(r.get('avg_grade', 0))
        soft  = float(r.get('soft_skills') or 0)
        hard  = float(r.get('hard_skills') or 0)
        ojt   = float(r.get('ojt_grade')   or 0)
        board = float(r.get('board_passer') or 0)
        return round(min(avg_g * 0.35 + ojt * 0.20 + soft * 0.15 + hard * 0.15 + board * 15, 100), 1)

    def _tier(high_score, fast_months):
        """2×2 matrix → 3 tiers.
           high+fast → Likely Employable
           high+slow or low+fast → Employable
           low+slow  → Least Employable
        """
        if high_score and fast_months:         return 'Likely Employable'
        if not high_score and not fast_months: return 'Least Employable'
        return 'Employable'

    def _rf_prob(r):
        """Run RF model and return probability_employed (0-1), or None on failure."""
        try:
            result = ml_predictor.predict_details({
                'avg_grade':      float(r['avg_grade']      or 0),
                'avg_prof_grade': float(r['avg_prof_grade'] or 0),
                'avg_elec_grade': float(r['avg_elec_grade'] or 0),
                'ojt_grade':      float(r['ojt_grade']      or 0),
                'soft_skills':    float(r['soft_skills']    or 0),
                'hard_skills':    float(r['hard_skills']    or 0),
                'age':            int(r['age']              or 22),
                'graduation_year': int(r['graduation_year'] or 2023),
                'course':         r['course'] or '',
            }, model='rf')
            p = result.get('probability_employed')
            return float(p) if p is not None else None
        except Exception:
            return None

    def _employability_level(r, pred_months=None, rf_probability=None):
        keys      = r.keys() if hasattr(r, 'keys') else []
        soft      = float(r['soft_skills'] or 0)
        hard      = float(r['hard_skills'] or 0)
        ncae_flag = bool(r['ncae_completed']) if 'ncae_completed' in keys else False
        if not ncae_flag and soft == 0 and hard == 0:
            return 'Pending Assessment'

        score     = _employability_score(r)
        employed  = bool(r['employed'])
        months    = r['months_to_employment'] if 'months_to_employment' in keys else None
        grad_year = r['graduation_year']       if 'graduation_year'       in keys else None

        rf_p = rf_probability if rf_probability is not None else 0.5

        # Latest-year students (graduating/new alumni) — use RF + score + predicted months
        if LATEST_YEAR and grad_year == LATEST_YEAR:
            high_s = score >= SCORE_THRESHOLD and rf_p >= 0.5
            fast_m = pred_months is not None and pred_months <= FAST_MONTHS
            return _tier(high_s, fast_m)

        # Historical employed alumni — use actual months + score
        if employed:
            high_s = score >= SCORE_THRESHOLD
            m      = int(months) if months is not None else None
            if m is None:
                return 'Likely Employable' if high_s else 'Employable'
            return _tier(high_s, m <= FAST_MONTHS)

        # Historical unemployed — RF probability + score
        high_s = score >= SCORE_THRESHOLD and rf_p >= 0.5
        return 'Likely Employable' if high_s else 'Least Employable'

    # Build lightweight user list (no ML yet — just plain fields)
    all_users = []
    for r in all_rows:
        employed  = bool(r.get('employed', 0))
        grad_year = r.get('graduation_year', 0)
        name = (r.get('name') or f"{r.get('first_name','')} {r.get('last_name','')}".strip()
                or f"Alumni {r.get('source_row_id','')}")
        all_users.append({
            '_raw':      r,
            'id':        r['id'],
            'name':      name,
            'email':     r.get('email', ''),
            'course':    r.get('course', 'Unknown'),
            'year':      grad_year,
            'status':    r.get('account_status', 'Active'),
            'employed':  employed,
            'board_passer':     bool(r.get('board_passer', 0)),
            'board_exam_score': float(r.get('board_exam_score') or 0.0),
            'months_to_employment': r.get('months_to_employment'),
        })

    # Compute employability stats directly from DB (score formula, no ML needed)
    # Match _employability_level logic exactly:
    # employed + months known: use _tier(score>=80, months<=4)
    # employed + months NULL:  Likely if score>=80 else Employable
    # unemployed:              Likely if score>=80 else Least
    _s = "(avg_grade*0.35 + ojt_grade*0.20 + soft_skills*0.15 + hard_skills*0.15 + board_passer*15)"
    emp_counts = db.execute(f"""
        SELECT
            SUM(CASE
                WHEN employed=1 AND months_to_employment IS NOT NULL AND {_s}>=75 AND months_to_employment<=6 THEN 1
                WHEN employed=1 AND months_to_employment IS NULL     AND {_s}>=75 THEN 1
                WHEN employed=0                                      AND {_s}>=75 THEN 1
                ELSE 0 END) AS likely,
            SUM(CASE
                WHEN employed=1 AND months_to_employment IS NOT NULL AND {_s}>=75 AND months_to_employment>6 THEN 1
                WHEN employed=1 AND months_to_employment IS NOT NULL AND {_s}<75  AND months_to_employment<=6 THEN 1
                WHEN employed=1 AND months_to_employment IS NULL     AND {_s}<75  THEN 1
                ELSE 0 END) AS employable,
            SUM(CASE
                WHEN employed=1 AND months_to_employment IS NOT NULL AND {_s}<75 AND months_to_employment>6 THEN 1
                WHEN employed=0                                      AND {_s}<75 THEN 1
                ELSE 0 END) AS least
        FROM ml_training_rows WHERE is_active = 1
    """).fetchone()

    stats = {
        'total':      len(all_users),
        'active':     sum(1 for u in all_users if u['status'] == 'Active'),
        'employed':   sum(1 for u in all_users if u['employed']),
        'unemployed': sum(1 for u in all_users if not u['employed']),
        'likely':     emp_counts['likely'] or 0,
        'employable': emp_counts['employable'] or 0,
        'least':      emp_counts['least'] or 0,
    }

    # Sort before pagination
    sort_by  = request.args.get('sort_by', '')
    sort_dir = request.args.get('sort_dir', 'asc')
    reverse  = sort_dir == 'desc'
    if sort_by == 'course':
        all_users.sort(key=lambda u: (u['course'] or '').lower(), reverse=reverse)
    elif sort_by == 'year':
        all_users.sort(key=lambda u: u['year'] or 0, reverse=reverse)
    elif sort_by == 'employment':
        all_users.sort(key=lambda u: 0 if u['employed'] else 1, reverse=reverse)
    elif sort_by == 'employability':
        all_users.sort(
            key=lambda u: _employability_score(u['_raw']),
            reverse=(sort_dir == 'desc')
        )

    # Filter
    processed = all_users
    if filter_by == 'Active':
        processed = [u for u in processed if u['status'] == 'Active']
    elif filter_by == 'Employed':
        processed = [u for u in processed if u['employed']]
    elif filter_by == 'Unemployed':
        processed = [u for u in processed if not u['employed']]
    if year_filter:
        try:
            yf = int(year_filter)
            processed = [u for u in processed if u['year'] == yf]
        except ValueError:
            pass
    if course_filter:
        processed = [u for u in processed if (u['course'] or '').upper() == course_filter.upper()]
    if search:
        processed = [u for u in processed if
                     search in u['name'].lower() or (u['email'] and search in u['email'].lower())]

    total_filtered = len(processed)

    # Paginate FIRST, then run ML only on the visible page
    try:
        page  = max(1, int(request.args.get('page', 1)))
        limit = max(1, int(request.args.get('limit', 100)))
    except (TypeError, ValueError):
        page, limit = 1, 100

    start = (page - 1) * limit
    paged_raw = processed[start:start + limit]

    # Run ML predictions on all paged rows (batch call — fast)
    raw_rows   = [u.pop('_raw') for u in paged_raw]
    batch_inputs = [{
        'avg_grade':      float(r.get('avg_grade') or 0),
        'avg_prof_grade': float(r.get('avg_prof_grade') or 0),
        'avg_elec_grade': float(r.get('avg_elec_grade') or 0),
        'ojt_grade':      float(r.get('ojt_grade') or 0),
        'soft_skills':    float(r.get('soft_skills') or 0),
        'hard_skills':    float(r.get('hard_skills') or 0),
        'age':            int(r.get('age') or 22),
        'graduation_year': int(r.get('graduation_year') or 0),
        'course':         r.get('course') or '',
    } for r in raw_rows]
    batch_preds = ml_predictor.predict_batch(batch_inputs, model='rf')

    paged_users = []
    for u, r, pred in zip(paged_raw, raw_rows, batch_preds):
        try:
            # All latest-year students are "graduating" regardless of employment
            is_grad = bool(LATEST_YEAR and u['year'] == LATEST_YEAR)
            rf_probability = pred.get('probability_employed')
            pred_months    = _knn_months(r) if is_grad and not u['employed'] else None
            score = _employability_score(r)
            level = _employability_level(r, pred_months, rf_probability)
        except Exception:
            is_grad = pred_months = rf_probability = None
            score = 0
            level = 'Pending Assessment'
        u.update({
            'predicted_months':    pred_months,
            'rf_probability':      round(rf_probability, 3) if rf_probability is not None else None,
            'is_graduating':       is_grad,
            'employability_score': score,
            'employability_level': level,
        })
        paged_users.append(u)

    return jsonify({
        'users': paged_users,
        'stats': stats,
        'latest_year': LATEST_YEAR,
        'available_years': available_years,
        'available_courses': available_courses,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': total_filtered,
            'pages': (total_filtered + limit - 1) // limit
        }
    }), 200


@admin_bp.route('/predict', methods=['GET'])
@admin_required
def predict_graduating():
    """Predict employability tiers for a selected year's students."""
    db = get_db()
    search      = request.args.get('search', '').lower()
    filter_tier = request.args.get('tier', 'All')

    # All available years
    year_rows = db.execute(
        "SELECT DISTINCT graduation_year FROM ml_training_rows WHERE is_active=1 ORDER BY graduation_year DESC"
    ).fetchall()
    available_years = [r['graduation_year'] for r in year_rows]

    if not available_years:
        return jsonify({'error': 'No dataset uploaded. Please upload a dataset first.'}), 404

    # Default to most recent year; allow override via ?year=
    latest = available_years[0]
    try:
        LATEST_YEAR = int(request.args.get('year', latest))
    except (TypeError, ValueError):
        LATEST_YEAR = latest

    if LATEST_YEAR not in available_years:
        LATEST_YEAR = latest

    SCORE_THRESHOLD = 75
    FAST_MONTHS = 6

    hist_rows = db.execute("""
        SELECT course, avg_grade, soft_skills, hard_skills,
               ojt_grade, avg_prof_grade, avg_elec_grade, months_to_employment
        FROM ml_training_rows
        WHERE is_active=1 AND employed=1 AND months_to_employment IS NOT NULL
    """).fetchall()

    # Previous-year alumni as reference pool for cross-year k-NN
    # Strictly previous years only (< not !=) — fallback to same-year if no history
    prev_rows = db.execute("""
        SELECT course, avg_grade, soft_skills, hard_skills,
               ojt_grade, avg_prof_grade, avg_elec_grade, employed
        FROM ml_training_rows
        WHERE is_active=1 AND graduation_year < ?
    """, [LATEST_YEAR]).fetchall()

    has_history = len(prev_rows) > 0

    if not has_history:
        prev_rows = db.execute("""
            SELECT course, avg_grade, soft_skills, hard_skills,
                   ojt_grade, avg_prof_grade, avg_elec_grade, employed
            FROM ml_training_rows WHERE is_active=1
        """).fetchall()

    _prev_by_course = {}
    for h in prev_rows:
        c = (h['course'] or '').upper().strip()
        _prev_by_course.setdefault(c, []).append(h)

    def _norm_gwa(g):
        g = float(g or 0)
        return round((5.0 - g) / 4.0 * 100, 1) if 0 < g <= 5.0 else g

    FEAT_COLS = ['avg_grade', 'soft_skills', 'hard_skills', 'ojt_grade', 'avg_prof_grade', 'avg_elec_grade']

    def _feats(r):
        return [_norm_gwa(r.get('avg_grade', 0)),
                float(r.get('soft_skills') or 0), float(r.get('hard_skills') or 0),
                float(r.get('ojt_grade') or 0), float(r.get('avg_prof_grade') or 0),
                float(r.get('avg_elec_grade') or 0)]

    def _knn_emp_rate(r, k=10):
        """Employment rate among k nearest previous-year alumni (same course preferred)."""
        course = (r.get('course') or '').upper().strip()
        pool = _prev_by_course.get(course) or list(prev_rows)
        if not pool:
            return None
        if len(pool) < k:
            pool = list(prev_rows)
        uf = _feats(r)
        neighbors = sorted(
            ((sum((a - float(h[col] or 0))**2 for a, col in zip(uf, FEAT_COLS))**0.5, int(h['employed']))
             for h in pool),
            key=lambda x: x[0]
        )[:k]
        if not neighbors:
            return None
        return sum(emp for _, emp in neighbors) / len(neighbors)

    def _score(r):
        avg_g = _norm_gwa(r.get('avg_grade', 0))
        return round(min(
            avg_g * 0.35 + float(r.get('ojt_grade') or 0) * 0.20 +
            float(r.get('soft_skills') or 0) * 0.15 + float(r.get('hard_skills') or 0) * 0.15 +
            float(r.get('board_passer') or 0) * 15, 100
        ), 1)

    def _estimate_months(rf_prob, knn_emp_rate):
        """Estimate months-to-employment from RF probability and k-NN employment rate."""
        combined = ((rf_prob or 0.5) * 0.6 + (knn_emp_rate or 0.5) * 0.4)
        if combined >= 0.80: return round(1 + (1 - combined) * 5)   # 1-2 months
        if combined >= 0.65: return round(2 + (1 - combined) * 10)  # 2-4 months
        if combined >= 0.50: return round(4 + (1 - combined) * 12)  # 4-7 months
        return round(7 + (0.5 - combined) * 16)                     # 7-15 months

    def _tier(score, rf_prob, knn_emp_rate):
        """
        Likely Employable  : score >= 72 AND (RF >= 0.55 OR kNN emp rate >= 0.60)
        Employable         : score >= 58 OR RF >= 0.50 OR kNN emp rate >= 0.50
        Least Employable   : everything else
        """
        knn = knn_emp_rate if knn_emp_rate is not None else 0.5
        rf  = rf_prob      if rf_prob      is not None else 0.5
        if score >= 72 and (rf >= 0.55 or knn >= 0.60):
            return 'Likely Employable'
        if score >= 58 or rf >= 0.50 or knn >= 0.50:
            return 'Employable'
        return 'Least Employable'

    reg_rows = db.execute(
        "SELECT * FROM users WHERE role='alumni' AND graduation_year=? AND (employed=0 OR employed IS NULL)",
        [LATEST_YEAR]
    ).fetchall()
    reg_emails = {r['email'].lower() for r in reg_rows if r['email']}

    ds_rows = db.execute(
        "SELECT * FROM ml_training_rows WHERE is_active=1 AND graduation_year=? AND employed=0",
        [LATEST_YEAR]
    ).fetchall()

    all_students = []
    for r in reg_rows:
        d = dict(r)
        d['_source'] = 'registered'
        d['name'] = f"{d.get('first_name','')} {d.get('last_name','')}".strip()
        all_students.append(d)

    for r in ds_rows:
        email = (r['email'] or '').lower()
        if email and email in reg_emails:
            continue
        all_students.append({
            'id': f"ds_{r['id']}", '_source': 'dataset',
            'name': r['name'] or f"Student {r['source_row_id']}",
            'email': r['email'] or '', 'course': r['course'],
            'graduation_year': r['graduation_year'], 'employed': 0,
            'avg_grade': r['avg_grade'], 'avg_prof_grade': r['avg_prof_grade'],
            'avg_elec_grade': r['avg_elec_grade'], 'ojt_grade': r['ojt_grade'],
            'soft_skills': r['soft_skills'], 'hard_skills': r['hard_skills'],
            'board_passer': r['board_passer'], 'age': r['age'],
        })

    # Batch all RF predictions in one vectorized call
    batch_inputs = [{
        'avg_grade':      float(r.get('avg_grade') or 0),
        'avg_prof_grade': float(r.get('avg_prof_grade') or 0),
        'avg_elec_grade': float(r.get('avg_elec_grade') or 0),
        'ojt_grade':      float(r.get('ojt_grade') or 0),
        'soft_skills':    float(r.get('soft_skills') or 0),
        'hard_skills':    float(r.get('hard_skills') or 0),
        'age':            int(r.get('age') or 22),
        'graduation_year': int(r.get('graduation_year') or LATEST_YEAR),
        'course':         r.get('course') or '',
    } for r in all_students]

    batch_preds = ml_predictor.predict_batch(batch_inputs, model='rf')

    all_results = []
    for r, pred in zip(all_students, batch_preds):
        try:
            rf_prob      = float(pred.get('probability_employed') or 0.5)
            knn_emp_rate = _knn_emp_rate(r)
            score        = _score(r)
            tier         = _tier(score, rf_prob, knn_emp_rate)
            est_months   = _estimate_months(rf_prob, knn_emp_rate)
            all_results.append({
                'id': r['id'], 'name': r.get('name') or 'Unknown',
                'email': r.get('email', ''), 'course': r.get('course', ''),
                'graduation_year': LATEST_YEAR, 'score': score,
                'rf_probability': round(rf_prob * 100, 1),
                'knn_emp_rate': round((knn_emp_rate or 0) * 100, 1),
                'predicted_months': est_months,
                'tier': tier,
                'board_passer': bool(r.get('board_passer', 0)),
                'source': r.get('_source', 'registered'),
            })
        except Exception as e:
            print(f"[predict] Error: {e}")
            continue

    all_results.sort(key=lambda x: x['score'], reverse=True)

    summary = {
        'high':       sum(1 for r in all_results if r['tier'] == 'Likely Employable'),
        'employable': sum(1 for r in all_results if r['tier'] == 'Employable'),
        'least':      sum(1 for r in all_results if r['tier'] == 'Least Employable'),
        'total':      len(all_results),
    }

    filtered = all_results
    if filter_tier != 'All':
        filtered = [r for r in filtered if r['tier'] == filter_tier]
    if search:
        filtered = [r for r in filtered if search in r['name'].lower() or search in r['course'].lower()]

    return jsonify({
        'students': filtered,
        'summary': summary,
        'graduation_year': LATEST_YEAR,
        'available_years': available_years,
    }), 200


@admin_bp.route('/predict/insights/<path:student_id>', methods=['GET'])
@admin_required
def predict_insights(student_id):
    """Return score breakdown + k-NN historical matches for a student."""
    db = get_db()

    # Fetch the student row
    if str(student_id).startswith('ds_'):
        rid = str(student_id).replace('ds_', '')
        row = db.execute('SELECT * FROM ml_training_rows WHERE id = ?', [rid]).fetchone()
        if not row:
            return jsonify({'error': 'Student not found'}), 404
        r = dict(row)
    else:
        row = db.execute('SELECT * FROM users WHERE id = ?', [student_id]).fetchone()
        if not row:
            return jsonify({'error': 'Student not found'}), 404
        r = dict(row)

    def _norm_gwa(g):
        g = float(g or 0)
        return round((5.0 - g) / 4.0 * 100, 1) if 0 < g <= 5.0 else g

    avg_g  = _norm_gwa(r.get('avg_grade', 0))
    ojt    = float(r.get('ojt_grade') or 0)
    soft   = float(r.get('soft_skills') or 0)
    hard   = float(r.get('hard_skills') or 0)
    board  = float(r.get('board_passer') or 0)
    prof   = float(r.get('avg_prof_grade') or 0)
    elec   = float(r.get('avg_elec_grade') or 0)
    score  = round(min(avg_g*0.35 + ojt*0.20 + soft*0.15 + hard*0.15 + board*15, 100), 1)

    score_breakdown = [
        {'label': 'Academic Grade',    'value': round(avg_g, 1),  'weighted': round(avg_g*0.35, 1),  'weight': '35%'},
        {'label': 'OJT / Internship',  'value': round(ojt, 1),   'weighted': round(ojt*0.20, 1),   'weight': '20%'},
        {'label': 'Soft Skills',       'value': round(soft, 1),  'weighted': round(soft*0.15, 1),  'weight': '15%'},
        {'label': 'Hard Skills',       'value': round(hard, 1),  'weighted': round(hard*0.15, 1),  'weight': '15%'},
        {'label': 'Board/Licensure',   'value': int(board),      'weighted': round(board*15, 1),   'weight': 'bonus'},
    ]

    # Get latest year for reference
    latest_yr = db.execute(
        'SELECT MAX(graduation_year) FROM ml_training_rows WHERE is_active=1'
    ).fetchone()[0]

    # Use strictly PREVIOUS years (< student's year), fall back to same-year if no history
    student_year = r.get('graduation_year') or latest_yr
    prev_rows = db.execute("""
        SELECT name, course, graduation_year, avg_grade, soft_skills, hard_skills,
               ojt_grade, avg_prof_grade, avg_elec_grade, employed, months_to_employment
        FROM ml_training_rows
        WHERE is_active=1 AND graduation_year < ?
    """, [student_year]).fetchall()

    using_fallback = len(prev_rows) == 0
    if using_fallback:
        prev_rows = db.execute("""
            SELECT name, course, graduation_year, avg_grade, soft_skills, hard_skills,
                   ojt_grade, avg_prof_grade, avg_elec_grade, employed, months_to_employment
            FROM ml_training_rows WHERE is_active=1 AND graduation_year = ?
        """, [student_year]).fetchall()

    FEAT_COLS = ['avg_grade', 'soft_skills', 'hard_skills', 'ojt_grade', 'avg_prof_grade', 'avg_elec_grade']
    uf = [avg_g, soft, hard, ojt, prof, elec]
    course = (r.get('course') or '').upper().strip()

    neighbors = []
    for h in prev_rows:
        dist = sum((a - float(h[col] or 0))**2 for a, col in zip(uf, FEAT_COLS)) ** 0.5
        neighbors.append((dist, h))
    neighbors.sort(key=lambda x: x[0])

    # Top 10 — prefer same course
    same = [(d, h) for d, h in neighbors if (h['course'] or '').upper().strip() == course][:6]
    other = [(d, h) for d, h in neighbors if (h['course'] or '').upper().strip() != course][:4]
    top10 = (same + other)[:10]

    matches = [{
        'name':        h['name'] or f"Alumni {i+1}",
        'course':      h['course'],
        'year':        h['graduation_year'],
        'avg_grade':   round(float(h['avg_grade'] or 0), 1),
        'soft_skills': round(float(h['soft_skills'] or 0), 1),
        'hard_skills': round(float(h['hard_skills'] or 0), 1),
        'ojt_grade':   round(float(h['ojt_grade'] or 0), 1),
        'employed':    bool(h['employed']),
        'months':      h['months_to_employment'],
        'similarity':  round(100 - min(d * 2, 99), 1),
    } for i, (d, h) in enumerate(top10)]

    emp_rate = round(sum(1 for m in matches if m['employed']) / len(matches) * 100, 1) if matches else 0

    return jsonify({
        'student': {
            'name':   r.get('name') or f"{r.get('first_name','')} {r.get('last_name','')}".strip(),
            'course': r.get('course', ''),
            'year':   r.get('graduation_year'),
            'score':  score,
        },
        'score_breakdown':    score_breakdown,
        'knn_matches':        matches,
        'knn_emp_rate':       emp_rate,
        'using_same_year':    using_fallback,
    }), 200


@admin_bp.route('/fill-pending-skills', methods=['POST'])
@admin_required
def fill_pending_skills():
    """Fill hard_skills/soft_skills for alumni who haven't completed NCAE using dataset averages."""
    db = get_db()
    pending = db.execute("""
        SELECT u.id, u.course, u.graduation_year
        FROM users u
        WHERE u.role = 'alumni'
          AND (u.ncae_completed = 0 OR u.ncae_completed IS NULL)
          AND (u.soft_skills = 0 OR u.soft_skills IS NULL)
          AND (u.hard_skills = 0 OR u.hard_skills IS NULL)
    """).fetchall()

    filled = 0
    for pu in pending:
        uid, course, grad_year = pu['id'], pu['course'], pu['graduation_year']
        row = db.execute("""
            SELECT AVG(soft_skills) AS avg_soft, AVG(hard_skills) AS avg_hard
            FROM ml_training_rows
            WHERE is_active = 1 AND course = ? AND graduation_year = ?
        """, [course, grad_year]).fetchone()

        avg_soft = row['avg_soft'] if row else None
        avg_hard = row['avg_hard'] if row else None

        if avg_soft is None:
            row = db.execute("""
                SELECT AVG(soft_skills) AS avg_soft, AVG(hard_skills) AS avg_hard
                FROM ml_training_rows WHERE is_active = 1 AND course = ?
            """, [course]).fetchone()
            avg_soft = row['avg_soft'] if row else None
            avg_hard = row['avg_hard'] if row else None

        if avg_soft is not None and avg_hard is not None:
            db.execute(
                "UPDATE users SET soft_skills = ?, hard_skills = ? WHERE id = ?",
                [round(float(avg_soft), 2), round(float(avg_hard), 2), uid]
            )
            filled += 1

    db.commit()
    return jsonify({'filled': filled, 'pending_total': len(pending)}), 200


@admin_bp.route('/users/ds_<int:ds_id>', methods=['GET', 'PUT'])
@admin_required
def manage_dataset_user(ds_id):
    db = get_db()
    if request.method == 'GET':
        r = db.execute('SELECT * FROM ml_training_rows WHERE id = ?', [ds_id]).fetchone()
        if not r:
            return jsonify({'error': 'User not found'}), 404
        return jsonify({'user': {
            'id':               f'ds_{r["id"]}',
            'firstName':        (r['name'] or '').split(',')[1].strip().split()[0] if ',' in (r['name'] or '') else (r['name'] or '').split()[0] if r['name'] else '',
            'middleName':       '',
            'lastName':         (r['name'] or '').split(',')[0].strip() if ',' in (r['name'] or '') else '',
            'email':            r['email'] or '',
            'course':           r['course'] or '',
            'graduationYear':   r['graduation_year'],
            'age':              r['age'],
            'employed':         bool(r['employed']),
            'monthsToEmployment': r['months_to_employment'],
            'avgGrade':         r['avg_grade'],
            'avgProfGrade':     r['avg_prof_grade'],
            'avgElecGrade':     r['avg_elec_grade'],
            'ojtGrade':         r['ojt_grade'],
            'softSkills':       r['soft_skills'],
            'hardSkills':       r['hard_skills'],
            'boardPasser':      bool(r['board_passer']),
            'boardExamScore':   r['board_exam_score'],
            'status':           'Active',
        }}), 200

    data = request.get_json()
    allowed_ds = {
        'employed':           ('employed',               lambda v: int(bool(v))),
        'monthsToEmployment': ('months_to_employment',   lambda v: int(v) if v not in (None, '') else None),
        'avgGrade':           ('avg_grade',              float),
        'avgProfGrade':       ('avg_prof_grade',         float),
        'avgElecGrade':       ('avg_elec_grade',         float),
        'ojtGrade':           ('ojt_grade',              float),
        'softSkills':         ('soft_skills',            float),
        'hardSkills':         ('hard_skills',            float),
        'boardPasser':        ('board_passer',           lambda v: int(bool(v))),
        'boardExamScore':     ('board_exam_score',       float),
        'course':             ('course',                 str),
        'graduationYear':     ('graduation_year',        int),
    }
    sets, vals = [], []
    for key, (col, cast) in allowed_ds.items():
        if key in data:
            try:
                val = cast(data[key]) if data[key] not in (None, '') else None
            except (ValueError, TypeError):
                val = None
            sets.append(f'{col} = ?')
            vals.append(val)
    if sets:
        vals.append(ds_id)
        db.execute(f"UPDATE ml_training_rows SET {', '.join(sets)} WHERE id = ?", vals)
        db.commit()
    return jsonify({'message': 'Updated'}), 200


@admin_bp.route('/users/<int:user_id>', methods=['GET', 'PUT'])
@admin_required
def manage_user(user_id):
    db = get_db()

    if request.method == 'GET':
        u = db.execute('SELECT * FROM users WHERE id = ?', [user_id]).fetchone()
        if not u:
            return jsonify({'error': 'User not found'}), 404
        return jsonify({'user': {
            'id': u['id'],
            'firstName': u['first_name'],
            'middleName': u['middle_name'] or '',
            'lastName': u['last_name'],
            'email': u['email'],
            'course': u['course'] or '',
            'graduationYear': u['graduation_year'],
            'age': u['age'],
            'employed': bool(u['employed']),
            'monthsToEmployment': u['months_to_employment'],
            'avgGrade': u['avg_grade'],
            'avgProfGrade': u['avg_prof_grade'],
            'avgElecGrade': u['avg_elec_grade'],
            'ojtGrade': u['ojt_grade'],
            'softSkills': u['soft_skills'],
            'hardSkills': u['hard_skills'],
            'boardPasser': bool(u['board_passer']),
            'boardExamScore': u['board_exam_score'],
            'status': u['account_status'],
        }}), 200

    # PUT
    data = request.get_json()

    allowed = {
        'firstName':          ('first_name',             str),
        'middleName':         ('middle_name',             str),
        'lastName':           ('last_name',              str),
        'email':              ('email',                  str),
        'course':             ('course',                 str),
        'graduationYear':     ('graduation_year',        int),
        'age':                ('age',                    int),
        'employed':           ('employed',               lambda v: int(bool(v))),
        'monthsToEmployment': ('months_to_employment',   lambda v: int(v) if v not in (None, '') else None),
        'avgGrade':           ('avg_grade',              float),
        'avgProfGrade':       ('avg_prof_grade',         float),
        'avgElecGrade':       ('avg_elec_grade',         float),
        'ojtGrade':           ('ojt_grade',              float),
        'softSkills':         ('soft_skills',            float),
        'hardSkills':         ('hard_skills',            float),
        'boardPasser':        ('board_passer',           lambda v: int(bool(v))),
        'boardExamScore':     ('board_exam_score',       float),
        'status':             ('account_status',         str),
    }

    sets, vals = [], []
    for key, (col, cast) in allowed.items():
        if key in data:
            try:
                val = cast(data[key]) if data[key] not in (None, '') else None
            except (ValueError, TypeError):
                val = None
            sets.append(f'{col} = ?')
            vals.append(val)

    if sets:
        vals.append(user_id)
        db.execute(f"UPDATE users SET {', '.join(sets)} WHERE id = ?", vals)
        db.commit()

    return jsonify({'message': 'User updated'}), 200


@admin_bp.route('/user-insights/<user_id>', methods=['GET'])
@admin_required
def user_insights(user_id):
    db = get_db()
    u = None
    user_id_str = str(user_id)
    
    print(f"[DEBUG] Fetching insights for ID: {user_id_str}")
    
    if user_id_str.startswith('ds_'):
        actual_id = user_id_str.replace('ds_', '')
        u = db.execute('SELECT * FROM ml_training_rows WHERE id = ?', [actual_id]).fetchone()
        if u:
            u = dict(u)
            # Add virtual fields to match 'users' table structure
            u['account_status'] = 'Dataset'
            u['first_name'] = u.get('name') or f"Alumni {u.get('source_row_id','')}"
            u['last_name'] = ''
            print(f"[DEBUG] Found in ml_training_rows: {u['first_name']}")
    else:
        u = db.execute('SELECT * FROM users WHERE id = ?', [user_id_str]).fetchone()
        if u:
            u = dict(u)
            print(f"[DEBUG] Found in users table: {u['first_name']} {u['last_name']}")

    if not u:
        print(f"[DEBUG] User not found: {user_id_str}")
        return jsonify({'error': f'User {user_id_str} not found'}), 404

    course = u.get('course')
    SCORE_THRESHOLD = 75
    FAST_MONTHS     = 6

    lr = db.execute(
        "SELECT MAX(graduation_year) AS my FROM ml_training_rows WHERE is_active=1"
    ).fetchone()
    LATEST_YEAR = lr['my'] if (lr and lr['my']) else None

    def _norm_gwa(g):
        g = float(g or 0)
        return round((5.0 - g) / 4.0 * 100, 1) if 0 < g <= 5.0 else g

    avg_g  = _norm_gwa(u['avg_grade'])
    soft   = float(u['soft_skills']    or 0)
    hard   = float(u['hard_skills']    or 0)
    ojt    = float(u['ojt_grade']      or 0)
    prof   = float(u['avg_prof_grade'] or 0)
    elec   = float(u['avg_elec_grade'] or 0)
    board  = float(u['board_passer']   or 0)
    score  = round(min(avg_g*0.35 + ojt*0.20 + soft*0.15 + hard*0.15 + board*15, 100), 1)

    user_feats = [avg_g, soft, hard, ojt, prof, elec]
    feat_cols  = ['avg_grade','soft_skills','hard_skills','ojt_grade','avg_prof_grade','avg_elec_grade']

    is_graduating = (LATEST_YEAR and u['graduation_year'] == LATEST_YEAR and not bool(u['employed']))

    # k-NN: find 10 most similar historical employed alumni
    course = (u['course'] or '').upper().strip()
    hist_pool = db.execute("""
        SELECT source_row_id, course, graduation_year, employed, months_to_employment,
               avg_grade, soft_skills, hard_skills, ojt_grade, avg_prof_grade, avg_elec_grade
        FROM ml_training_rows
        WHERE is_active=1 AND employed=1 AND months_to_employment IS NOT NULL
          AND course = ?
    """, [course]).fetchall()
    if len(hist_pool) < 10:
        hist_pool = db.execute("""
            SELECT source_row_id, course, graduation_year, employed, months_to_employment,
                   avg_grade, soft_skills, hard_skills, ojt_grade, avg_prof_grade, avg_elec_grade
            FROM ml_training_rows
            WHERE is_active=1 AND employed=1 AND months_to_employment IS NOT NULL
        """).fetchall()

    dists = sorted(
        [(sum((a - float(h[c] or 0))**2 for a, c in zip(user_feats, feat_cols))**0.5, h)
         for h in hist_pool],
        key=lambda x: x[0]
    )[:5]

    similar = []
    total_w = sum(1.0/(d+1e-6) for d, _ in dists)
    pred_months = round(sum(int(h['months_to_employment'])/(d+1e-6) for d, h in dists) / total_w) if dists else None

    for dist, h in dists:
        similar.append({
            'alumni_id': h['source_row_id'] or '—',
            'course':    h['course'],
            'year':      h['graduation_year'],
            'employed':  bool(h['employed']),
            'months':    int(h['months_to_employment']) if h['months_to_employment'] else None,
            'distance':  round(dist, 2),
        })

    # RF probability
    rf_prob = None
    try:
        u_age = u['age'] if ('age' in u.keys() and u['age'] is not None) else 22
        res = ml_predictor.predict_details({
            'avg_grade': float(u['avg_grade'] or 0), 'avg_prof_grade': float(u['avg_prof_grade'] or 0),
            'avg_elec_grade': float(u['avg_elec_grade'] or 0), 'ojt_grade': float(u['ojt_grade'] or 0),
            'soft_skills': float(u['soft_skills'] or 0), 'hard_skills': float(u['hard_skills'] or 0),
            'age': int(u_age), 'graduation_year': int(u['graduation_year'] or 2023),
            'course': u['course'] or '',
        }, model='rf')
        p = res.get('probability_employed')
        rf_prob = round(float(p), 3) if p is not None else None
    except Exception:
        pass

    # Peer Comparison & Strengths/Improvement (using the large dataset for accuracy)
    peer_averages = db.execute("""
        SELECT 
            AVG(avg_grade) as avg_grade,
            AVG(soft_skills) as soft_skills,
            AVG(hard_skills) as hard_skills,
            AVG(ojt_grade) as ojt_grade,
            AVG(avg_prof_grade) as avg_prof_grade,
            AVG(avg_elec_grade) as avg_elec_grade
        FROM ml_training_rows 
        WHERE is_active = 1 AND course = ?
    """, [course]).fetchone()

    peer_norm = {
        'avg_grade':      _norm_gwa(peer_averages['avg_grade']),
        'soft_skills':    float(peer_averages['soft_skills'] or 0),
        'hard_skills':    float(peer_averages['hard_skills'] or 0),
        'ojt_grade':      float(peer_averages['ojt_grade'] or 0),
        'avg_prof_grade': float(peer_averages['avg_prof_grade'] or 0),
        'avg_elec_grade': float(peer_averages['avg_elec_grade'] or 0),
    }

    strengths = []
    improvements = []
    
    metrics_map = {
        'GPA / Grade': (avg_g, peer_norm['avg_grade']),
        'Soft Skills': (soft, peer_norm['soft_skills']),
        'Hard Skills': (hard, peer_norm['hard_skills']),
        'OJT Performance': (ojt, peer_norm['ojt_grade']),
        'Professional Subjects': (prof, peer_norm['avg_prof_grade']),
    }

    for label, (val, peer_val) in metrics_map.items():
        diff = val - peer_val
        if diff > 5:
            strengths.append({'label': label, 'impact': 'high', 'diff': round(diff, 1)})
        elif diff < -10:
            improvements.append({'label': label, 'impact': 'medium', 'diff': round(diff, 1)})

    return jsonify({
        'score': score,
        'score_breakdown': {
            'avg_grade':      round(avg_g, 1),
            'soft_skills':    round(soft, 1),
            'hard_skills':    round(hard, 1),
            'ojt_grade':      round(ojt, 1),
            'avg_prof_grade': round(prof, 1),
            'avg_elec_grade': round(elec, 1),
        },
        'peer_comparison': peer_norm,
        'strengths': strengths,
        'improvements': improvements,
        'predicted_months': pred_months,
        'rf_probability':   rf_prob,
        'is_graduating':    is_graduating,
        'employability_level': 'N/A',
        'similar_alumni':   similar,
    }), 200


# ── Forecasting ────────────────────────────────────────────────────────────

@admin_bp.route('/forecasting', methods=['GET'])
@admin_required
def get_forecasting():
    db = get_db()

    _ensure_employment_data_from_training(db)
    emp_rows = db.execute(
        "SELECT year, overall_rate, male_rate, female_rate FROM employment_data ORDER BY year"
    ).fetchall()

    historical = [{'year': str(r['year']), 'rate': r['overall_rate']} for r in emp_rows]
    rates = [r['overall_rate'] for r in emp_rows]
    years = [r['year'] for r in emp_rows]
    if not years:
        return jsonify({
            'historical_data': [],
            'forecast_data': [],
            'course_data': [],
            'projected_values': [],
            'model_metrics': {},
            'model_used': 'Linear Regression',
            'message': 'No employment data available.',
        }), 200

    # Default 3-year forecast
    result = _forecast_result_for_model(rates, horizon=3, model_str='Linear Regression')
    forecast_points = []
    for i, val in enumerate(result['forecast_values']):
        forecast_points.append({
            'year': str(max(years) + i + 1),
            'rate': val,
            'forecast': True,
        })

    by_course = _get_program_rates(db, max(years))
    
    # Fetch full names mapping
    p_rows = db.execute("SELECT code, name FROM programs").fetchall()
    p_map = {r['code']: r['name'] for r in p_rows}
    
    course_data = [
        {
            'course': p_map.get(r['course'], r['course']), 
            'rate': r['rate']
        } for r in by_course
    ]

    return jsonify({
        'historical_data': historical,
        'forecast_data': historical + forecast_points,
        'course_data': course_data,
        'projected_values': [
            {'year': str(max(years) + i + 1), 'val': f"{v}%"}
            for i, v in enumerate(result['forecast_values'])
        ],
        'model_metrics': result['metrics'],
        'model_used': result.get('model_used', 'Linear Regression'),
    }), 200


@admin_bp.route('/forecasting/run', methods=['POST'])
@admin_required
def run_forecasting():
    data = request.get_json()
    horizon = int(data.get('horizon', 3))
    model_str = data.get('model', 'Linear Regression')

    db = get_db()
    _ensure_employment_data_from_training(db)
    emp_rows = db.execute(
        "SELECT year, overall_rate FROM employment_data ORDER BY year"
    ).fetchall()

    rates = [r['overall_rate'] for r in emp_rows]
    years = [r['year'] for r in emp_rows]
    if not years:
        return jsonify({
            'data': [],
            'forecast_values': [],
            'metrics': {},
            'model_used': model_str,
            'message': 'No employment data available.',
        }), 200

    result = _forecast_result_for_model(rates, horizon=horizon, model_str=model_str)

    historical = [{'year': str(r['year']), 'rate': r['overall_rate']} for r in emp_rows]
    forecast_points = []
    for i, val in enumerate(result['forecast_values']):
        forecast_points.append({
            'year': str(max(years) + i + 1),
            'rate': val,
            'forecast': True,
        })

    return jsonify({
        'data': historical + forecast_points,
        'forecast_values': [
            {'year': str(max(years) + i + 1), 'val': f"{v}%"}
            for i, v in enumerate(result['forecast_values'])
        ],
        'metrics': result['metrics'],
        'model_used': result.get('model_used', model_str),
    }), 200


# ── Employment Comparison ──────────────────────────────────────────────────

@admin_bp.route('/employment-comparison', methods=['GET'])
@admin_required
def employment_comparison():
    db = get_db()

    _ensure_employment_data_from_training(db)
    rows = db.execute(
        "SELECT year, overall_rate, male_rate, female_rate FROM employment_data ORDER BY year"
    ).fetchall()
    if not rows:
        return jsonify({
            'by_year': [],
            'by_course': [],
            'by_gender': [],
            'summary': {
                'avg_rate': 'N/A',
                'avg_delta': '',
                'best_prog': 'N/A',
                'best_rate': '',
                'peak_year': 'N/A',
                'peak_rate': '',
                'gender_gap': 'N/A',
                'gender_note': '',
            },
        }), 200

    by_year = [{
        'year': str(r['year']),
        'employed': r['overall_rate'],
        'unemployed': round(100 - r['overall_rate'], 1),
    } for r in rows]

    by_gender = [{
        'year': str(r['year']),
        'male': r['male_rate'],
        'female': r['female_rate'],
    } for r in rows]

    latest_year = max(r['year'] for r in rows)
    course_rows = _get_program_rates(db, latest_year)
    by_course = [{'course': r['course'], 'rate': r['rate']} for r in course_rows]

    rates = [r['overall_rate'] for r in rows]
    avg_rate = round(sum(rates) / len(rates), 1) if rates else 0
    first_rate = rates[0] if rates else 0
    last_rate = rates[-1] if rates else 0
    change = round(last_rate - first_rate, 1)

    best_course = by_course[0]['course'] if by_course else 'BSCS'
    best_rate = by_course[0]['rate'] if by_course else 82
    peak_year = str(max(rows, key=lambda r: r['overall_rate'])['year'])
    peak_rate = max(r['overall_rate'] for r in rows)

    gender_gap = round(
        sum(r['male_rate'] - r['female_rate'] for r in rows) / len(rows), 1
    ) if rows else 3.0

    return jsonify({
        'by_year': by_year,
        'by_course': by_course,
        'by_gender': by_gender,
        'summary': {
            'avg_rate': f'{avg_rate}%',
            'avg_delta': f'+{change}%',
            'best_prog': best_course,
            'best_rate': f'{best_rate}% rate',
            'peak_year': peak_year,
            'peak_rate': f'{peak_rate}%',
            'gender_gap': f'{gender_gap}%',
            'gender_note': 'Male higher',
        },
    }), 200


# ── Predict & Report ───────────────────────────────────────────────────────

@admin_bp.route('/predict-employability', methods=['POST'])
@admin_required
def predict_employability_route():
    payload = request.get_json() or {}
    db = get_db()

    features = payload.get('features')
    user_id = payload.get('user_id')
    requested_model = _normalize_model_choice(payload.get('model', 'rf'))

    if features is None and user_id is not None:
        features = _alumni_features_from_db(db, user_id)
        if not features:
            return jsonify({'error': 'Alumni user not found'}), 404

    if not isinstance(features, dict):
        return jsonify({'error': 'Provide a features object or user_id'}), 400

    settings = _read_prediction_settings(db)
    use_voter_weights = settings['use_voter_weights']

    if requested_model == 'arima':
        ml_details = _predict_with_arima_employability(db, features)
    else:
        ml_details = predict_employability_details(features, model=requested_model)
    voter_fields = _read_voter_fields(db)
    voter_details = _predict_with_voter_weights(features, voter_fields)

    mode = 'ml_default'
    details = ml_details
    if use_voter_weights:
        mode = 'voter_weighted'
        details = voter_details
    elif ml_details.get('error'):
        mode = 'voter_fallback'
        details = voter_details

    if details.get('error'):
        return jsonify({'error': details['error']}), 503

    return jsonify({
        'prediction': {
            **details,
            'mode': mode,
            'use_voter_weights': use_voter_weights,
            'requested_model': requested_model,
            'ml_reference': None if ml_details.get('error') else ml_details,
        }
    }), 200


@admin_bp.route('/models/status', methods=['GET'])
@admin_required
def model_status():
    return jsonify({'model': predictor_status()}), 200


TEST_ACCOUNT_EMAILS = (
    'juan.delacruz@plp.edu.ph', 'maria.santos@plp.edu.ph', 'pedro.reyes@plp.edu.ph',
    'ana.garcia@plp.edu.ph', 'jose.mendoza@plp.edu.ph',
    # demo accounts (ncae_completed=0)
    'demo.alumni@plp.edu.ph', 'demo2.alumni@plp.edu.ph',
    # preview accounts — 2 per program (ncae_completed=1)
    'bscs.1@plp.edu.ph', 'bscs.2@plp.edu.ph',
    'bsit.1@plp.edu.ph', 'bsit.2@plp.edu.ph',
    'bscpe.1@plp.edu.ph', 'bscpe.2@plp.edu.ph',
    'bsece.1@plp.edu.ph', 'bsece.2@plp.edu.ph',
    'bsce.1@plp.edu.ph', 'bsce.2@plp.edu.ph',
    'bsn.1@plp.edu.ph', 'bsn.2@plp.edu.ph',
    'bsed.1@plp.edu.ph', 'bsed.2@plp.edu.ph',
    'beed.1@plp.edu.ph', 'beed.2@plp.edu.ph',
    'bsa.1@plp.edu.ph', 'bsa.2@plp.edu.ph',
    'bsba.1@plp.edu.ph', 'bsba.2@plp.edu.ph',
    'bshm.1@plp.edu.ph', 'bshm.2@plp.edu.ph',
)

def _sync_users_to_training_rows(db):
    """Copy alumni user records into ml_training_rows so models can train."""
    ep = ','.join('?' * len(TEST_ACCOUNT_EMAILS))
    alumni = db.execute(f"""
        SELECT id, first_name, last_name, email, course, graduation_year, age,
               avg_grade, avg_prof_grade, avg_elec_grade, ojt_grade,
               soft_skills, hard_skills, board_passer, board_exam_score, employed
        FROM users
        WHERE role = 'alumni' AND graduation_year IS NOT NULL
          AND (is_test_account = 0 OR is_test_account IS NULL)
          AND LOWER(email) NOT IN ({ep})
    """, list(TEST_ACCOUNT_EMAILS)).fetchall()

    inserted = 0
    for u in alumni:
        source_row_id = f"user_{u['id']}"
        name = f"{u['first_name'] or ''} {u['last_name'] or ''}".strip()
        try:
            db.execute("""
                INSERT INTO ml_training_rows (
                    source_name, source_row_id, name, email, course,
                    graduation_year, age, avg_grade, avg_prof_grade,
                    avg_elec_grade, ojt_grade, soft_skills, hard_skills,
                    board_passer, board_exam_score, employed, is_active, imported_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,datetime('now'))
                ON CONFLICT(source_name, source_row_id) DO UPDATE SET
                    avg_grade=excluded.avg_grade,
                    avg_prof_grade=excluded.avg_prof_grade,
                    avg_elec_grade=excluded.avg_elec_grade,
                    ojt_grade=excluded.ojt_grade,
                    soft_skills=excluded.soft_skills,
                    hard_skills=excluded.hard_skills,
                    employed=excluded.employed,
                    is_active=1
            """, [
                'users_sync', source_row_id, name, u['email'] or '',
                u['course'] or '', int(u['graduation_year']),
                int(u['age'] or 22),
                float(u['avg_grade'] or 0), float(u['avg_prof_grade'] or 0),
                float(u['avg_elec_grade'] or 0), float(u['ojt_grade'] or 0),
                float(u['soft_skills'] or 0), float(u['hard_skills'] or 0),
                int(u['board_passer'] or 0), float(u['board_exam_score'] or 0),
                int(u['employed'] or 0),
            ])
            inserted += 1
        except Exception:
            continue
    db.commit()
    return inserted


@admin_bp.route('/models/sync-and-retrain', methods=['POST'])
@admin_required
def sync_and_retrain():
    """Sync users → ml_training_rows then retrain models. Used when background import failed."""
    db = get_db()
    db_path = current_app.config.get('DATABASE', os.getenv('DATABASE', 'plp_alumni.db'))

    # Remove test account data from training rows (in case they were previously synced)
    ep = ','.join('?' * len(TEST_ACCOUNT_EMAILS))
    db.execute(f"DELETE FROM ml_training_rows WHERE LOWER(email) IN ({ep}) OR source_name='users_sync'",
               list(TEST_ACCOUNT_EMAILS))
    db.commit()
    # Clear employment_data and program_rates if no real training data remains
    real_count = db.execute("SELECT COUNT(*) FROM ml_training_rows WHERE is_active=1").fetchone()[0]
    if real_count == 0:
        db.execute("DELETE FROM employment_data")
        db.execute("DELETE FROM program_rates")
        db.commit()
    # Remove accounts for years with no training data
    _sync_alumni_to_training_data(db)
    synced = _sync_users_to_training_rows(db)
    total_rows = db.execute(
        "SELECT COUNT(*) FROM ml_training_rows WHERE is_active=1"
    ).fetchone()[0]

    if total_rows == 0:
        return jsonify({'error': 'No alumni data found to train on.'}), 400

    rf_metadata = train_random_forest(database_path=db_path)
    lr_metadata = train_linear_employability(database_path=db_path)
    ml_predictor._load_models()
    return jsonify({
        'message': f'Synced {synced} users and retrained models.',
        'synced_rows': synced,
        'total_training_rows': total_rows,
        'models': {'rf': rf_metadata, 'lr': lr_metadata},
    }), 200


@admin_bp.route('/models/retrain', methods=['POST'])
@admin_required
def retrain_model():
    db = get_db()
    db_path = current_app.config.get('DATABASE', os.getenv('DATABASE', 'plp_alumni.db'))

    # Auto-sync from users if training rows are empty
    total_rows = db.execute(
        "SELECT COUNT(*) FROM ml_training_rows WHERE is_active=1"
    ).fetchone()[0]
    if total_rows == 0:
        _sync_users_to_training_rows(db)

    rf_metadata = train_random_forest(database_path=db_path)
    lr_metadata = train_linear_employability(database_path=db_path)
    ml_predictor._load_models()
    return jsonify({
        'message': 'Model retrained successfully',
        'model': rf_metadata,
        'models': {'rf': rf_metadata, 'lr': lr_metadata},
    }), 200


@admin_bp.route('/models/import-first-clean-dataset', methods=['POST'])
@admin_required
def import_first_clean_dataset():
    payload = request.get_json() or {}
    retrain_after_import = bool(payload.get('retrain_after_import', True))
    db_path = current_app.config.get('DATABASE', os.getenv('DATABASE', 'plp_alumni.db'))
    dataset_name = 'first_clean_dataset.csv'
    dataset_path = os.path.join(current_app.root_path, 'ml', 'data', dataset_name)

    try:
        imported = import_first_clean_dataset_rows(
            database_path=db_path,
            csv_path=dataset_path,
            source_name=dataset_name,
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    response = {
        'message': 'Dataset imported into ML training rows',
        'import': imported,
    }

    if retrain_after_import:
        rf_metadata = train_random_forest(database_path=db_path)
        lr_metadata = train_linear_employability(database_path=db_path)
        ml_predictor._load_models()
        response['message'] = 'Dataset imported and model retrained'
        response['models'] = {
            'rf': rf_metadata,
            'lr': lr_metadata,
        }

    return jsonify(response), 200


@admin_bp.route('/predict-report', methods=['GET'])
@admin_required
def predict_report():
    db = get_db()

    reports = db.execute(
        "SELECT * FROM reports ORDER BY created_at DESC"
    ).fetchall()

    report_list = [{
        'id': r['id'],
        'name': r['name'],
        'date': r['created_at'][:10] if r['created_at'] else '',
        'type': r['type'],
        'status': r['status'],
        'year_range': r['year_range'],
    } for r in reports]

    # Run all forecast models to get latest metrics
    emp_rows = db.execute(
        "SELECT overall_rate FROM employment_data ORDER BY year"
    ).fetchall()
    rates = [r['overall_rate'] for r in emp_rows]
    model_runs = {
        'Linear Regression': _forecast_result_for_model(rates, horizon=1, model_str='Linear Regression'),
        'Random Forest Regressor': _forecast_result_for_model(rates, horizon=1, model_str='Random Forest'),
        'Auto ARIMA (AIC search)': _forecast_result_for_model(rates, horizon=1, model_str='Auto ARIMA (AIC search)'),
    }
    default_metrics = model_runs['Linear Regression']['metrics']
    metrics_by_model = {
        model_name: {
            'mae': str(result['metrics']['mae']),
            'rmse': str(result['metrics']['rmse']),
            'mape': _format_percent_metric(result['metrics']['mape']),
            'r2': str(result['metrics']['r2']),
        }
        for model_name, result in model_runs.items()
    }

    return jsonify({
        'reports': report_list,
        'metrics': {
            'mae': str(default_metrics['mae']),
            'rmse': str(default_metrics['rmse']),
            'mape': _format_percent_metric(default_metrics['mape']),
            'r2': str(default_metrics['r2']),
        },
        'metrics_by_model': metrics_by_model,
    }), 200


@admin_bp.route('/predict-report/generate', methods=['POST'])
@admin_required
def generate_report():
    data = request.get_json()
    report_type = data.get('type', data.get('report_type', 'PDF'))
    year_range = data.get('year_range', '2019\u20132024')
    model_name = data.get('model', 'Linear Regression')

    report_name = f"Employment Forecast Report ({year_range})"
    db = get_db()
    cur = db.execute("""
        INSERT INTO reports (name, type, year_range, model_name, status)
        VALUES (?,?,?,?,?)
    """, [report_name, report_type, year_range, model_name, 'Ready'])
    db.commit()

    # Get metrics
    emp_rows = db.execute("SELECT overall_rate FROM employment_data ORDER BY year").fetchall()
    rates = [r['overall_rate'] for r in emp_rows]
    metrics_by_model = None
    model_choice = _normalize_model_choice(model_name)
    if str(model_name or '').strip().lower() == 'all models':
        model_runs = {
            'Linear Regression': _forecast_result_for_model(rates, horizon=1, model_str='Linear Regression'),
            'Random Forest Regressor': _forecast_result_for_model(rates, horizon=1, model_str='Random Forest'),
            'Auto ARIMA (AIC search)': _forecast_result_for_model(rates, horizon=1, model_str='Auto ARIMA (AIC search)'),
        }
        fm = model_runs['Linear Regression']['metrics']
        metrics_by_model = {
            model: {
                'mae': str(result['metrics']['mae']),
                'rmse': str(result['metrics']['rmse']),
                'mape': _format_percent_metric(result['metrics']['mape']),
                'r2': str(result['metrics']['r2']),
            }
            for model, result in model_runs.items()
        }
    elif model_choice == 'lr':
        fm = _forecast_result_for_model(rates, horizon=1, model_str='Linear Regression')['metrics']
    elif model_choice == 'rf':
        fm = _forecast_result_for_model(rates, horizon=1, model_str='Random Forest')['metrics']
    else:
        fm = _forecast_result_for_model(rates, horizon=1, model_str=model_name)['metrics']

    from datetime import date
    return jsonify({
        'message': 'Report generated successfully',
        'report': {
            'id': cur.lastrowid,
            'name': report_name,
            'date': date.today().strftime('%b %d, %Y'),
            'type': report_type,
            'status': 'Ready',
        },
        'metrics': {
            'mae': str(fm['mae']),
            'rmse': str(fm['rmse']),
            'mape': _format_percent_metric(fm['mape']),
            'r2': str(fm['r2']),
        },
        'metrics_by_model': metrics_by_model,
    }), 201


@admin_bp.route('/reports/count', methods=['GET'])
@admin_required
def report_row_count():
    db = get_db()
    years_raw = request.args.getlist('years')
    programs_raw = request.args.getlist('programs')
    query = "SELECT COUNT(*) as cnt FROM ml_training_rows WHERE is_active=1"
    params = []
    if years_raw:
        placeholders = ','.join('?' * len(years_raw))
        query += f" AND graduation_year IN ({placeholders})"
        params += [int(y) for y in years_raw]
    if programs_raw:
        placeholders = ','.join('?' * len(programs_raw))
        query += f" AND UPPER(TRIM(course)) IN ({placeholders})"
        params += [p.upper().strip() for p in programs_raw]
    count = db.execute(query, params).fetchone()['cnt']
    return jsonify({'count': count}), 200


@admin_bp.route('/reports/download', methods=['GET'])
@admin_required
def download_report():
    import io
    from flask import make_response

    fmt = request.args.get('format', 'excel').lower()
    years_raw = request.args.getlist('years')
    programs_raw = request.args.getlist('programs')

    db = get_db()

    # Fetch alumni with prediction scores
    query = "SELECT * FROM ml_training_rows WHERE is_active=1"
    params = []
    if years_raw:
        placeholders = ','.join('?' * len(years_raw))
        query += f" AND graduation_year IN ({placeholders})"
        params += [int(y) for y in years_raw]
    if programs_raw:
        placeholders = ','.join('?' * len(programs_raw))
        query += f" AND UPPER(TRIM(course)) IN ({placeholders})"
        params += [p.upper().strip() for p in programs_raw]
    query += " ORDER BY graduation_year, course, name"
    rows = db.execute(query, params).fetchall()

    # Compute employability score + tier for each row
    def _score(r):
        gwa = float(r['avg_grade'] or 0)
        norm = round((5.0 - gwa) / 4.0 * 100, 1) if 0 < gwa <= 5.0 else gwa
        ojt   = float(r['ojt_grade']   or 0)
        soft  = float(r['soft_skills'] or 0)
        hard  = float(r['hard_skills'] or 0)
        board = float(r['board_passer'] or 0)
        return round(min(norm * 0.35 + ojt * 0.20 + soft * 0.15 + hard * 0.15 + board * 15, 100), 1)

    def _tier(score, employed):
        if score >= 80: return 'Likely Employable'
        if score >= 65: return 'Employable'
        return 'Least Employable'

    enriched = []
    for r in rows:
        sc = _score(r)
        tier = _tier(sc, r['employed'])
        enriched.append({
            'name': r['name'] or '',
            'email': r['email'] or '',
            'course': r['course'] or '',
            'graduation_year': r['graduation_year'],
            'employment_status': 'Employed' if r['employed'] else 'Unemployed',
            'employability_score': sc,
            'employability_tier': tier,
            'gwa': round(float(r['avg_grade'] or 0), 2),
            'soft_skills': round(float(r['soft_skills'] or 0), 1),
            'hard_skills': round(float(r['hard_skills'] or 0), 1),
            'board_passer': 'Yes' if r['board_passer'] else 'No',
        })

    years_label = ', '.join(sorted(set(str(r['graduation_year']) for r in rows))) or 'All Years'

    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from collections import defaultdict

        wb = openpyxl.Workbook()
        hdr_fill = PatternFill('solid', fgColor='163D22')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        tot_fill = PatternFill('solid', fgColor='E6EDE8')
        tot_font = Font(bold=True, color='163D22', size=10)
        center  = Alignment(horizontal='center')

        # ── Sheet 1: Summary by Year & Program ───────────────────────────
        ws1 = wb.active
        ws1.title = 'Summary'
        ws1['A1'] = 'PLP Alumni Employability Prediction Report'
        ws1['A1'].font = Font(bold=True, size=14, color='163D22')
        ws1['A2'] = f'Years: {years_label}   |   Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws1['A2'].font = Font(size=10, color='666666')
        ws1.append([])

        heads = ['Year', 'Program', 'Total', 'Likely Employable', 'Employable', 'Least Employable', 'Employed', 'Unemployed', 'Employment Rate']
        ws1.append(heads)
        for c, h in enumerate(heads, 1):
            cell = ws1.cell(row=4, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center

        agg = defaultdict(lambda: {'total':0,'likely':0,'emp':0,'least':0,'employed':0})
        for e in enriched:
            key = (e['graduation_year'], e['course'])
            agg[key]['total'] += 1
            if e['employability_tier'] == 'Likely Employable': agg[key]['likely'] += 1
            elif e['employability_tier'] == 'Employable': agg[key]['emp'] += 1
            else: agg[key]['least'] += 1
            if e['employment_status'] == 'Employed': agg[key]['employed'] += 1

        rn = 5
        for (yr, prog), v in sorted(agg.items()):
            t = v['total']; em = v['employed']
            rate = f"{round(em/t*100,1)}%" if t else '0%'
            unemp = t - em
            ws1.append([yr, prog, t, v['likely'], v['emp'], v['least'], em, unemp, rate])
            for c in range(1, 10): ws1.cell(row=rn, column=c).alignment = center
            rn += 1

        tot = len(enriched)
        tot_emp = sum(1 for e in enriched if e['employment_status'] == 'Employed')
        tot_likely = sum(1 for e in enriched if e['employability_tier'] == 'Likely Employable')
        tot_employable = sum(1 for e in enriched if e['employability_tier'] == 'Employable')
        tot_least = sum(1 for e in enriched if e['employability_tier'] == 'Least Employable')
        ws1.append(['TOTAL', 'All', tot, tot_likely, tot_employable, tot_least, tot_emp, tot-tot_emp,
                    f"{round(tot_emp/tot*100,1)}%" if tot else '0%'])
        for c in range(1, 10):
            cell = ws1.cell(row=rn, column=c)
            cell.fill = tot_fill; cell.font = tot_font; cell.alignment = center

        col_widths = [8, 12, 8, 16, 12, 16, 10, 12, 16]
        for i, w in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        # ── Sheet 2: Alumni with Predictions ─────────────────────────────
        ws2 = wb.create_sheet('Alumni Predictions')
        det_heads = ['Name', 'Email', 'Program', 'Year', 'Score', 'Employability Tier',
                     'Employment Status', 'GWA', 'Soft Skills', 'Hard Skills', 'Board Passer']
        ws2.append(det_heads)
        for c, h in enumerate(det_heads, 1):
            cell = ws2.cell(row=1, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center

        tier_colors = {'Likely Employable': 'D1FAE5', 'Employable': 'DBEAFE', 'Least Employable': 'FEE2E2'}
        for i, e in enumerate(enriched, 2):
            row = [e['name'], e['email'], e['course'], e['graduation_year'],
                   e['employability_score'], e['employability_tier'], e['employment_status'],
                   e['gwa'], e['soft_skills'], e['hard_skills'], e['board_passer']]
            ws2.append(row)
            tier_fill = PatternFill('solid', fgColor=tier_colors.get(e['employability_tier'], 'FFFFFF'))
            ws2.cell(row=i, column=6).fill = tier_fill

        det_widths = [22, 26, 10, 8, 8, 18, 18, 8, 12, 12, 14]
        for i, w in enumerate(det_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        resp = make_response(output.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename=PLP_Employability_Report_{years_label.replace(", ","_")}.xlsx'
        return resp

    else:  # PDF via printable HTML
        from collections import defaultdict
        tot = len(enriched)
        tot_emp = sum(1 for e in enriched if e['employment_status'] == 'Employed')
        tot_likely = sum(1 for e in enriched if e['employability_tier'] == 'Likely Employable')
        tot_employable = sum(1 for e in enriched if e['employability_tier'] == 'Employable')
        tot_least = sum(1 for e in enriched if e['employability_tier'] == 'Least Employable')
        emp_rate = round(tot_emp / tot * 100, 1) if tot else 0

        agg = defaultdict(lambda: {'total':0,'likely':0,'emp':0,'least':0,'employed':0})
        for e in enriched:
            key = (e['graduation_year'], e['course'])
            agg[key]['total'] += 1
            if e['employability_tier'] == 'Likely Employable': agg[key]['likely'] += 1
            elif e['employability_tier'] == 'Employable': agg[key]['emp'] += 1
            else: agg[key]['least'] += 1
            if e['employment_status'] == 'Employed': agg[key]['employed'] += 1

        summary_rows = ''
        for (yr, prog), v in sorted(agg.items()):
            t = v['total']; em = v['employed']
            rate = f"{round(em/t*100,1)}%" if t else '0%'
            summary_rows += f'<tr><td>{yr}</td><td>{prog}</td><td>{t}</td><td class="likely">{v["likely"]}</td><td class="emp">{v["emp"]}</td><td class="least">{v["least"]}</td><td>{em}</td><td>{t-em}</td><td>{rate}</td></tr>'

        tier_badge = {'Likely Employable':'#15803d','Employable':'#1d4ed8','Least Employable':'#b91c1c'}
        tier_bg    = {'Likely Employable':'#dcfce7','Employable':'#dbeafe','Least Employable':'#fee2e2'}

        detail_rows = ''.join(
            f'<tr><td>{e["name"]}</td><td>{e["course"]}</td><td>{e["graduation_year"]}</td>'
            f'<td>{e["gwa"]}</td><td>{e["soft_skills"]}</td><td>{e["hard_skills"]}</td>'
            f'<td>{e["employability_score"]}</td>'
            f'<td><span style="background:{tier_bg.get(e["employability_tier"],"#f3f4f6")};color:{tier_badge.get(e["employability_tier"],"#333")};padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700">{e["employability_tier"]}</span></td>'
            f'<td>{e["employment_status"]}</td></tr>'
            for e in enriched[:500]  # limit to 500 rows for PDF readability
        )

        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>PLP Employability Report</title>
<style>
  body{{font-family:Arial,sans-serif;margin:24px;color:#111}}
  h1{{color:#163d22;font-size:20px;margin:0 0 4px}}
  .sub{{color:#666;font-size:11px;margin-bottom:16px}}
  .stats{{display:flex;gap:12px;margin:16px 0;flex-wrap:wrap}}
  .stat{{background:#e6ede8;border-radius:8px;padding:10px 16px;text-align:center;min-width:90px}}
  .stat-val{{font-size:22px;font-weight:900;color:#163d22}}
  .stat-lbl{{font-size:10px;color:#555;margin-top:1px}}
  .likely{{color:#15803d;font-weight:700}} .emp{{color:#1d4ed8;font-weight:700}} .least{{color:#b91c1c;font-weight:700}}
  h2{{font-size:13px;color:#163d22;margin:20px 0 8px;border-bottom:2px solid #e6ede8;padding-bottom:4px}}
  table{{width:100%;border-collapse:collapse;font-size:11px}}
  th{{background:#163d22;color:#fff;padding:6px 8px;text-align:left}}
  td{{padding:5px 8px;border-bottom:1px solid #f0f0f0}}
  tr:nth-child(even) td{{background:#fafafa}}
  .total-row td{{background:#e6ede8;font-weight:700}}
  @media print{{button{{display:none}}}}
</style></head><body>
<button onclick="window.print()" style="float:right;padding:6px 14px;background:#163d22;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px">🖨 Print / Save PDF</button>
<h1>PLP Alumni Employability Prediction Report</h1>
<div class="sub">Years: {years_label} &nbsp;|&nbsp; Generated: {datetime.now().strftime('%B %d, %Y')}</div>
<div class="stats">
  <div class="stat"><div class="stat-val">{tot:,}</div><div class="stat-lbl">Total Alumni</div></div>
  <div class="stat"><div class="stat-val likely">{tot_likely:,}</div><div class="stat-lbl">Likely Employable</div></div>
  <div class="stat"><div class="stat-val emp">{tot_employable:,}</div><div class="stat-lbl">Employable</div></div>
  <div class="stat"><div class="stat-val least">{tot_least:,}</div><div class="stat-lbl">Least Employable</div></div>
  <div class="stat"><div class="stat-val">{emp_rate}%</div><div class="stat-lbl">Employment Rate</div></div>
</div>
<h2>Summary by Year & Program</h2>
<table><tr><th>Year</th><th>Program</th><th>Total</th><th>Likely</th><th>Employable</th><th>Least</th><th>Employed</th><th>Unemployed</th><th>Rate</th></tr>
{summary_rows}
<tr class="total-row"><td colspan="2">TOTAL</td><td>{tot:,}</td><td class="likely">{tot_likely}</td><td class="emp">{tot_employable}</td><td class="least">{tot_least}</td><td>{tot_emp}</td><td>{tot-tot_emp}</td><td>{emp_rate}%</td></tr></table>
<h2>Alumni Predictions{' (first 500 shown)' if len(enriched) > 500 else ''}</h2>
<table><tr><th>Name</th><th>Program</th><th>Year</th><th>GWA</th><th>Soft</th><th>Hard</th><th>Score</th><th>Tier</th><th>Status</th></tr>
{detail_rows}</table>
</body></html>"""

        resp = make_response(html)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        return resp


# ── Voter Config ───────────────────────────────────────────────────────────

@admin_bp.route('/voter-config', methods=['GET'])
@admin_required
def get_voter_config():
    db = get_db()
    fields = _read_voter_fields(db)
    settings = _read_prediction_settings(db)
    return jsonify({
        'config': fields,
        'use_voter_weights': settings['use_voter_weights'],
    }), 200


@admin_bp.route('/voter-config', methods=['PUT'])
@admin_required
def update_voter_config():
    data = request.get_json() or {}
    fields = data.get('config', data.get('fields', []))
    use_voter_weights = bool(data.get('use_voter_weights', False))
    db = get_db()

    for field in fields:
        weight = int(field.get('weight', 0) or 0)
        weight = max(0, min(100, weight))
        db.execute("""
            UPDATE voter_config SET enabled = ?, weight = ? WHERE field_key = ?
        """, [int(field.get('enabled', True)), weight, field['key']])

    db.execute("""
        INSERT INTO prediction_settings (id, use_voter_weights, updated_at)
        VALUES (1, ?, datetime('now'))
        ON CONFLICT(id) DO UPDATE SET
            use_voter_weights = excluded.use_voter_weights,
            updated_at = excluded.updated_at
    """, [int(use_voter_weights)])

    db.commit()
    return jsonify({
        'message': 'Voter configuration saved',
        'use_voter_weights': use_voter_weights,
    }), 200


@admin_bp.route('/voter-config/suggest', methods=['POST'])
@admin_required
def suggest_voter_config():
    db = get_db()
    fields = _read_voter_fields(db)
    suggestion = _suggest_voter_weights_from_rf(fields)
    if suggestion.get('error'):
        return jsonify({'error': suggestion['error']}), 503
    return jsonify(suggestion), 200


@admin_bp.route('/predict-dataset', methods=['POST'])
@admin_required
def predict_dataset():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    # Save temp file
    temp_path = os.path.join('uploads', f"temp_predict_{file.filename}")
    os.makedirs('uploads', exist_ok=True)
    file.save(temp_path)

    try:
        from ml.dataset_importer import _normalize_columns, _map_row
        import pandas as pd
        from datetime import datetime

        # Load file
        ext = os.path.splitext(file.filename)[1].lower()
        if ext in ('.xlsx', '.xls'):
            df = pd.read_excel(temp_path)
        else:
            df = pd.read_csv(temp_path)

        raw_columns = list(df.columns)
        col_mapping = _normalize_columns(raw_columns)

        current_year = datetime.utcnow().year
        results = []

        for idx, row_series in df.iterrows():
            row_dict = {col_mapping.get(c, c): v for c, v in row_series.items()}
            # Note: _map_row expects some defaults or overrides
            mapped, err = _map_row(row_dict, str(idx+1), current_year, year_override=current_year)

            if mapped:
                # Run prediction
                pred = ml_predictor.predict_details({
                    'avg_grade':      mapped['avg_grade'],
                    'avg_prof_grade': mapped['avg_prof_grade'],
                    'avg_elec_grade': mapped['avg_elec_grade'],
                    'ojt_grade':      mapped['ojt_grade'],
                    'soft_skills':    mapped['soft_skills'],
                    'hard_skills':    mapped['hard_skills'],
                    'graduation_year': mapped['graduation_year'],
                    'course':         mapped['course'],
                    'board_passer':   mapped.get('board_passer', 0),
                    'board_exam_score': mapped.get('board_exam_score', 0),
                }, model='rf')

                results.append({
                    'name': mapped.get('name') or f"Candidate {idx+1}",
                    'course': mapped['course'],
                    'probability': round(pred.get('probability_employed', 0) * 100, 1),
                    'level': 'Likely Employable' if pred.get('probability_employed', 0) > 0.6 else 'Employable' if pred.get('probability_employed', 0) > 0.4 else 'Least Employable',
                    'key_factor': pred.get('top_feature', 'Academic Performance')
                })

        # Summary Stats
        total = len(results)
        likely = len([r for r in results if r['level'] == 'Likely Employable'])

        return jsonify({
            'filename': file.filename,
            'total_candidates': total,
            'likely_employable': likely,
            'success_rate': round((likely/total)*100, 1) if total > 0 else 0,
            'predictions': results
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


# ── Factors Configuration ──────────────────────────────────────────────────

BOARD_EXAM_PROGRAMS = {'BSN', 'BSECE', 'BSED', 'PSYCHOLOGY', 'BSA'}


@admin_bp.route('/factors-config', methods=['GET'])
@admin_required
def get_factors_config():
    """Return voter_config factors with program-aware board/licensure filtering."""
    db = get_db()
    program = (request.args.get('program') or '').strip().upper()
    rows = db.execute("SELECT * FROM voter_config ORDER BY id").fetchall()
    is_board_program = not program or program in BOARD_EXAM_PROGRAMS

    factors = []
    for r in rows:
        if r['field_key'] == 'gender':
            continue  # never show gender
        if r['field_key'] == 'board_passer' and not is_board_program:
            continue  # hide board factor for non-board programs
        factors.append({
            'key':     r['field_key'],
            'name':    r['field_name'],
            'enabled': bool(r['enabled']),
            'weight':  r['weight'],
            'is_board_factor': r['field_key'] == 'board_passer',
        })

    total_weight = sum(f['weight'] for f in factors if f['enabled']) or 1
    for f in factors:
        f['pct'] = round(f['weight'] / total_weight * 100, 1) if f['enabled'] else 0

    return jsonify({
        'factors': factors,
        'program': program or None,
        'is_board_program': is_board_program,
        'board_programs': sorted(BOARD_EXAM_PROGRAMS),
    }), 200


@admin_bp.route('/factors-config', methods=['PUT'])
@admin_required
def update_factors_config():
    data = request.get_json() or {}
    fields = data.get('factors', [])
    db = get_db()
    for f in fields:
        weight = max(0, min(100, int(f.get('weight', 0) or 0)))
        db.execute(
            "UPDATE voter_config SET enabled=?, weight=? WHERE field_key=?",
            [int(bool(f.get('enabled', True))), weight, f['key']]
        )
    db.commit()
    return jsonify({'message': 'Saved'}), 200


@admin_bp.route('/factors-configuration', methods=['GET'])
@admin_required
def factors_configuration():
    model_key = (request.args.get('model') or 'lr').strip().lower()
    program = (request.args.get('program') or '').strip()
    limit = request.args.get('limit', 10)
    try:
        limit = max(1, int(limit))
    except (TypeError, ValueError):
        limit = 10

    if program:
        try:
            mapping, err = _importance_from_program_data(model_key, program)
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
    else:
        mapping, err = _importance_from_loaded_model(model_key)

    if err:
        return jsonify({'error': err}), 503
    if not mapping:
        return jsonify({'error': 'No factor data available.'}), 503

    # Program-Aware Factor Filtering
    from ml.predictor import BOARD_PROGRAMS
    active_program = program.upper() if program else None
    
    # If a specific program is selected and it's NOT a board program, hide board factors
    if active_program and active_program not in BOARD_PROGRAMS:
        mapping = {k: v for k, v in mapping.items() if k not in ('board_passer', 'board_exam_score')}

    # Prefer actionable inputs; fill remaining slots with course-level signals if needed.
    primary = {k: v for k, v in mapping.items() if not k.startswith('course_')}
    primary_items = sorted(primary.items(), key=lambda kv: kv[1], reverse=True)
    items = primary_items[:limit]

    if len(items) < limit:
        course_items = sorted(
            ((k, v) for k, v in mapping.items() if k.startswith('course_')),
            key=lambda kv: kv[1],
            reverse=True,
        )
        for item in course_items:
            if len(items) >= limit:
                break
            items.append(item)
    total = sum(float(v) for _, v in items) or 1.0

    factors = []
    for feature, val in items:
        weight = round((float(val) / total) * 100, 2)
        factors.append({
            'feature': feature,
            'label': _format_factor_label(feature),
            'importance': round(float(val), 6),
            'weight': weight,
        })

    return jsonify({
        'model': model_key,
        'program': program.upper() if program else None,
        'factors': factors,
    }), 200


# ── Upload Model ───────────────────────────────────────────────────────────

def _auto_forecast_3yr(db, dataset_year):
    """Compute employment rate for dataset_year from training rows, update employment_data,
    then run all 3 models with horizon=3 and return structured forecast."""
    row = db.execute("""
        SELECT COUNT(*) as total, COALESCE(SUM(employed), 0) as emp
        FROM ml_training_rows WHERE graduation_year = ? AND is_active = 1
    """, [dataset_year]).fetchone()
    if row['total'] > 0:
        rate = round((row['emp'] / row['total']) * 100, 2)
        db.execute("""
            INSERT INTO employment_data (year, overall_rate, employed_count, unemployed_count)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(year) DO UPDATE SET
                overall_rate = excluded.overall_rate,
                employed_count = excluded.employed_count,
                unemployed_count = excluded.unemployed_count
        """, [dataset_year, rate, int(row['emp']), row['total'] - int(row['emp'])])
        db.commit()

    _refresh_program_rates(db, dataset_year)

    emp_rows = db.execute(
        "SELECT year, overall_rate FROM employment_data ORDER BY year"
    ).fetchall()
    if not emp_rows:
        return None

    rates = [r['overall_rate'] for r in emp_rows]
    base_year = max(r['year'] for r in emp_rows)
    forecast_years = [base_year + 1, base_year + 2, base_year + 3]
    predictions = {}
    for key, model_str in [('lr', 'Linear Regression'), ('rf', 'Random Forest'), ('arima', 'Auto ARIMA (AIC search)')]:
        try:
            result = _forecast_result_for_model(rates, horizon=3, model_str=model_str)
            fv = result.get('forecast_values', [])
            predictions[key] = [
                {'year': forecast_years[i], 'rate': round(float(v), 2)}
                for i, v in enumerate(fv) if i < 3
            ]
        except Exception:
            predictions[key] = []

    return {'base_year': dataset_year, 'forecast_years': forecast_years, 'predictions': predictions}


@admin_bp.route('/upload', methods=['POST'])
@admin_required
def upload_model():
    _set_progress('uploading', 0, 'Preparing upload…')
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    model_name = request.form.get('name', file.filename)
    overwrite_all = _is_truthy(request.form.get('overwrite_all'), default=False)
    apply_to_training = True if overwrite_all else _is_truthy(request.form.get('apply_to_training'), default=False)
    retrain_after_import = True if overwrite_all else _is_truthy(request.form.get('retrain_after_import'), default=True)
    dataset_year_raw = request.form.get('dataset_year', '').strip()
    conflict_mode = request.form.get('conflict_mode', '').lower()  # 'overwrite' or 'merge'

    dataset_year = None
    if dataset_year_raw:
        try:
            dataset_year = int(dataset_year_raw)
        except ValueError:
            return jsonify({'error': 'dataset_year must be a valid integer'}), 400

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
    os.makedirs(upload_folder, exist_ok=True)

    safe_name = file.filename.replace(' ', '_')
    file_path = os.path.join(upload_folder, safe_name)
    _set_progress('uploading', 2, 'Saving file…')
    file.save(file_path)
    file_size = os.path.getsize(file_path)
    file_hash = _file_sha256(file_path)
    _set_progress('uploading', 4, 'File saved')
    is_csv = safe_name.lower().endswith('.csv')

    is_tabular = is_csv or safe_name.lower().endswith('.xlsx') or safe_name.lower().endswith('.xls')
    if apply_to_training and not is_tabular:
        return jsonify({'error': 'Training import is only supported for CSV or Excel (.xlsx/.xls) files.'}), 400

    records = 0
    if is_tabular:
        try:
            if is_csv:
                import csv as _csv
                with open(file_path, newline='', encoding='utf-8-sig') as f:
                    records = max(sum(1 for _ in _csv.reader(f)) - 1, 0)
            else:
                import pandas as _pd
                records = len(_pd.read_excel(file_path))
        except Exception as exc:
            return jsonify({'error': f'Unable to parse file rows: {exc}'}), 400

    db = get_db()

    # ── Year sequence + conflict check (skipped for bulk overwrite) ──────────
    if not overwrite_all and apply_to_training and is_csv and dataset_year:
        max_year_row = db.execute(
            "SELECT MAX(graduation_year) AS my FROM ml_training_rows WHERE is_active = 1"
        ).fetchone()
        max_year = max_year_row['my'] if max_year_row else None

        # Reject year that skips ahead (e.g. 2022 exists, trying to upload 2024)
        if max_year and dataset_year > max_year + 1:
            return jsonify({
                'error': f'Cannot upload year {dataset_year}. '
                         f'Year {max_year + 1} is missing — upload that first.',
                'next_allowed': max_year + 1,
                'last_recorded': max_year,
            }), 400

        existing = db.execute(
            "SELECT COUNT(*) FROM ml_training_rows WHERE graduation_year = ? AND is_active = 1",
            [dataset_year]
        ).fetchone()[0]
        if existing > 0 and not conflict_mode:
            return jsonify({
                'year_conflict': True,
                'year': dataset_year,
                'existing_count': existing,
                'message': f'{existing} training rows already exist for {dataset_year}.',
            }), 409

        if conflict_mode == 'overwrite' and dataset_year:
            db.execute("DELETE FROM ml_training_rows WHERE graduation_year = ?", [dataset_year])
            db.commit()

    cur = db.execute("""
        INSERT INTO model_uploads (
            name, original_filename, file_size, records, status, sha256, applied_to_training
        )
        VALUES (?,?,?,?,?,?,?)
    """, [model_name, safe_name, file_size, records, 'Active', file_hash, 0])
    db.commit()

    import_result = None
    trained_models = None
    training_policy = 'archive_only'
    forecast = None

    if apply_to_training and is_tabular:
        db_path = current_app.config.get('DATABASE', os.getenv('DATABASE', 'plp_alumni.db'))
        create_accounts = _is_truthy(request.form.get('create_accounts'), default=False)
        skip_email = _is_truthy(request.form.get('skip_email'), default=True)

        # Run import (+ optional retrain) fully in background to avoid Render's 60s timeout
        global _training_job
        with _training_lock:
            _training_job = {'status': 'running', 'result': None, 'error': None}
        _set_progress('importing', 2, 'Starting import in background…')

        app_obj = current_app._get_current_object()
        t = threading.Thread(
            target=_run_import_and_training_background,
            args=(app_obj, db_path, file_path, safe_name,
                  dataset_year, retrain_after_import, cur.lastrowid,
                  create_accounts, skip_email, conflict_mode, overwrite_all),
            daemon=True,
        )
        t.start()
        training_policy = 'async_import_and_training'

    upload_row = db.execute(
        "SELECT * FROM model_uploads WHERE id = ?", [cur.lastrowid]
    ).fetchone()

    upload = {
        'id': upload_row['id'],
        'name': upload_row['name'],
        'filename': upload_row['original_filename'],
        'size': f"{upload_row['file_size'] / 1024:.2f} KB",
        'records': f"{upload_row['records']} records",
        'status': upload_row['status'],
        'date': datetime.now().strftime('%Y-%m-%d'),
        'sha256': upload_row['sha256'],
        'applied_to_training': bool(upload_row['applied_to_training']),
    }

    response = {
        'message': 'File uploaded successfully',
        'training_policy': training_policy,
        'upload': upload,
        'training_async': training_policy in ('uploaded_csv_imported_and_retraining_async', 'async_import_and_training'),
    }
    if import_result:
        response['import'] = import_result
    if trained_models:
        response['models'] = trained_models
        response['message'] = 'File uploaded, imported to training, and model retrained'
    if forecast:
        response['forecast'] = forecast

    return jsonify(response), 201


@admin_bp.route('/uploads', methods=['GET'])
@admin_required
def list_uploads():
    db = get_db()
    # Return only the latest upload per filename (deduped), ordered by date
    rows = db.execute("""
        SELECT * FROM model_uploads
        WHERE id IN (
            SELECT MAX(id) FROM model_uploads GROUP BY original_filename
        )
        ORDER BY uploaded_at DESC
    """).fetchall()

    uploads = [{
        'id': r['id'],
        'name': r['name'],
        'filename': r['original_filename'],
        'size': f"{r['file_size'] / 1024:.2f} KB",
        'records': f"{r['records']:,} rows" if isinstance(r['records'], int) else f"{r['records']} rows",
        'status': r['status'],
        'date': r['uploaded_at'][:10] if r['uploaded_at'] else '',
        'applied_to_training': bool(r['applied_to_training']),
    } for r in rows]

    return jsonify({'uploads': uploads}), 200


@admin_bp.route('/training/status', methods=['GET'])
@admin_required
def training_status():
    with _training_lock:
        return jsonify(dict(_training_job)), 200


@admin_bp.route('/upload/progress', methods=['GET'])
@admin_required
def upload_progress():
    with _upload_progress_lock:
        return jsonify(dict(_upload_progress)), 200


# ── All-Models Forecasting ─────────────────────────────────────────────────

@admin_bp.route('/forecasting/run-all', methods=['POST'])
@admin_required
def run_forecasting_all_models():
    """Run all 3 models and return combined chart data for merged graph."""
    data = request.get_json()
    horizon = int(data.get('horizon', 3))

    db = get_db()
    _ensure_employment_data_from_training(db)

    emp_rows = db.execute(
        "SELECT year, overall_rate FROM employment_data ORDER BY year"
    ).fetchall()
    rates = [r['overall_rate'] for r in emp_rows]
    years = [r['year'] for r in emp_rows]
    if not rates:
        return jsonify({
            'data': [],
            'historical': [],
            'projections': {},
            'metrics': {},
            'horizon': horizon,
            'message': 'No employment data available.',
        }), 200

    historical = [{'year': str(r['year']), 'rate': r['overall_rate']} for r in emp_rows]

    model_map = {
        'lr': 'Linear Regression',
        'rf': 'Random Forest',
        'arima': 'Auto ARIMA (AIC search)',
    }
    combined = {str(r['year']): {'year': str(r['year']), 'rate': r['overall_rate']} for r in emp_rows}
    projections = {}
    metrics_all = {}

    for key, model_str in model_map.items():
        result = _forecast_result_for_model(rates, horizon=horizon, model_str=model_str)
        fv = result.get('forecast_values', [])
        for i, val in enumerate(fv):
            yr = str(max(years) + i + 1)
            if yr not in combined:
                combined[yr] = {'year': yr, 'forecast': True}
            combined[yr][key] = val
        projections[key] = [
            {'year': str(max(years) + i + 1), 'val': f"{v}%"}
            for i, v in enumerate(fv)
        ]
        metrics_all[key] = result.get('metrics', {})

    chart_data = list(combined.values())
    chart_data.sort(key=lambda x: x['year'])

    return jsonify({
        'data': chart_data,
        'historical': historical,
        'projections': projections,
        'metrics': metrics_all,
        'horizon': horizon,
    }), 200


# ── Programs Management ────────────────────────────────────────────────────

BOARD_EXAM_PROGRAMS = {
    'BSCE', 'BSEE', 'BSME', 'BSECE', 'BSN', 'BSEd', 'BEEd', 'BSA', 'BSCPE', 'BSMA', 'BSPH',
}

DEFAULT_PROGRAMS = [
    {'name': 'Bachelor of Science in Computer Science', 'code': 'BSCS', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
    {'name': 'Bachelor of Science in Information Technology', 'code': 'BSIT', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
    {'name': 'Bachelor of Science in Computer Engineering', 'code': 'BSCPE', 'has_board_exam': 1, 'board_exam_name': 'Electronics Engineering Licensure Exam', 'description': 'Combines hardware and software engineering disciplines.'},
    {'name': 'Bachelor of Science in Electronics Engineering', 'code': 'BSECE', 'has_board_exam': 1, 'board_exam_name': 'Electronics Engineering Licensure Exam', 'description': ''},
    {'name': 'Bachelor of Science in Civil Engineering', 'code': 'BSCE', 'has_board_exam': 1, 'board_exam_name': 'Civil Engineering Licensure Exam', 'description': ''},
    {'name': 'Bachelor of Science in Nursing', 'code': 'BSN', 'has_board_exam': 1, 'board_exam_name': 'Nurse Licensure Examination', 'description': ''},
    {'name': 'Bachelor of Secondary Education', 'code': 'BSEd', 'has_board_exam': 1, 'board_exam_name': 'Licensure Examination for Teachers', 'description': ''},
    {'name': 'Bachelor of Elementary Education', 'code': 'BEEd', 'has_board_exam': 1, 'board_exam_name': 'Licensure Examination for Teachers', 'description': ''},
    {'name': 'Bachelor of Science in Accountancy', 'code': 'BSA', 'has_board_exam': 1, 'board_exam_name': 'CPA Licensure Examination', 'description': ''},
    {'name': 'Bachelor of Science in Business Administration', 'code': 'BSBA', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
    {'name': 'Bachelor of Science in Hotel and Restaurant Management', 'code': 'BSHM', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
    {'name': 'Bachelor of Arts in Psychology', 'code': 'PSYCHOLOGY', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
    {'name': 'Bachelor of Science in Entrepreneurship', 'code': 'BSENTREP', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
    {'name': 'Bachelor of Science in Information Systems', 'code': 'BSIS', 'has_board_exam': 0, 'board_exam_name': '', 'description': ''},
]

@admin_bp.route('/programs', methods=['GET'])
@admin_required
def list_programs():
    db = get_db()
    
    # 1. Ensure all DEFAULT_PROGRAMS exist with correct names
    for p in DEFAULT_PROGRAMS:
        exists = db.execute("SELECT id, name FROM programs WHERE code = ?", [p['code']]).fetchone()
        if not exists:
            db.execute("""
                INSERT INTO programs (name, code, has_board_exam, board_exam_name, description, status)
                VALUES (?,?,?,?,?,'Active')
            """, [p['name'], p['code'], p['has_board_exam'], p['board_exam_name'], p['description']])
        elif exists['name'] == p['code'] or not exists['name']:
            # Update legacy acronym-only names to full descriptive names
            db.execute("UPDATE programs SET name = ? WHERE code = ?", [p['name'], p['code']])
    
    # 2. Sync any additional programs found in ML Training Data
    ml_programs = db.execute("SELECT DISTINCT course FROM ml_training_rows WHERE is_active = 1").fetchall()
    for row in ml_programs:
        code = row['course']
        exists = db.execute("SELECT 1 FROM programs WHERE code = ?", [code]).fetchone()
        if not exists:
            name = code.replace('_', ' ').title()
            db.execute("""
                INSERT INTO programs (name, code, has_board_exam, board_exam_name, description, status)
                VALUES (?,?,?,?,'','Active')
            """, [name, code, 0, ''])
    
    db.commit()

    rows = db.execute('SELECT * FROM programs ORDER BY name').fetchall()
    programs_list = [{
        'id': r['id'],
        'name': r['name'],
        'code': r['code'],
        'has_board_exam': bool(r['has_board_exam']),
        'board_exam_name': r['board_exam_name'] or '',
        'description': r['description'] or '',
        'status': r['status'],
        'created_at': r['created_at'][:10] if r['created_at'] else '',
    } for r in rows]
    return jsonify({'programs': programs_list}), 200


@admin_bp.route('/programs', methods=['POST'])
@admin_required
def create_program():
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Program name is required'}), 400
    db = get_db()
    try:
        cur = db.execute("""
            INSERT INTO programs (name, code, has_board_exam, board_exam_name, description, status)
            VALUES (?,?,?,?,?,?)
        """, [
            name,
            (data.get('code') or '').strip().upper(),
            int(bool(data.get('has_board_exam', False))),
            (data.get('board_exam_name') or '').strip(),
            (data.get('description') or '').strip(),
            data.get('status', 'Active'),
        ])
        db.commit()
        return jsonify({'message': 'Program created', 'id': cur.lastrowid}), 201
    except Exception:
        return jsonify({'error': 'Program name already exists'}), 409


@admin_bp.route('/programs/<int:program_id>', methods=['PUT'])
@admin_required
def update_program(program_id):
    data = request.get_json()
    db = get_db()
    prog = db.execute('SELECT * FROM programs WHERE id = ?', [program_id]).fetchone()
    if not prog:
        return jsonify({'error': 'Program not found'}), 404
    try:
        db.execute("""
            UPDATE programs SET name=?, code=?, has_board_exam=?, board_exam_name=?, description=?, status=?
            WHERE id=?
        """, [
            (data.get('name') or prog['name']).strip(),
            (data.get('code') or prog['code'] or '').strip().upper(),
            int(bool(data.get('has_board_exam', bool(prog['has_board_exam'])))),
            (data.get('board_exam_name') or prog['board_exam_name'] or '').strip(),
            (data.get('description') or prog['description'] or '').strip(),
            data.get('status', prog['status']),
            program_id,
        ])
        db.commit()
        return jsonify({'message': 'Program updated'}), 200
    except Exception:
        return jsonify({'error': 'Program name already exists'}), 409


@admin_bp.route('/programs/<int:program_id>', methods=['DELETE'])
@admin_required
def delete_program(program_id):
    db = get_db()
    db.execute('DELETE FROM programs WHERE id = ?', [program_id])
    db.commit()
    return jsonify({'message': 'Program deleted'}), 200


# ── Company Account Management ─────────────────────────────────────────────

@admin_bp.route('/company-accounts', methods=['GET'])
@admin_required
def list_company_accounts():
    db = get_db()
    rows = db.execute("""
        SELECT u.id, u.first_name, u.last_name, u.email, u.account_status,
               u.company_id, c.name AS company_name, c.industry, u.created_at
        FROM users u
        LEFT JOIN companies c ON u.company_id = c.id
        WHERE u.role = 'company'
        ORDER BY u.created_at DESC
    """).fetchall()
    accounts = [{
        'id': r['id'],
        'name': f"{r['first_name']} {r['last_name']}",
        'email': r['email'],
        'status': r['account_status'],
        'company_id': r['company_id'],
        'company_name': r['company_name'] or '',
        'industry': r['industry'] or '',
        'created_at': r['created_at'][:10] if r['created_at'] else '',
    } for r in rows]
    return jsonify({'accounts': accounts}), 200


@admin_bp.route('/company-accounts', methods=['POST'])
@admin_required
def create_company_account():
    import bcrypt as _bcrypt
    data = request.get_json()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or 'company123'
    first_name = (data.get('first_name') or data.get('name') or 'Company').strip()
    last_name = (data.get('last_name') or 'User').strip()
    company_id = data.get('company_id')

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    db = get_db()
    existing = db.execute('SELECT id FROM users WHERE LOWER(email) = ?', [email]).fetchone()
    if existing:
        return jsonify({'error': 'Email already registered'}), 409

    pw_hash = _bcrypt.hashpw(password.encode(), _bcrypt.gensalt()).decode()
    cur = db.execute("""
        INSERT INTO users (first_name, last_name, email, password_hash, role, account_status, company_id)
        VALUES (?,?,?,?,'company','Active',?)
    """, [first_name, last_name, email, pw_hash, company_id])
    db.commit()
    return jsonify({'message': 'Company account created', 'id': cur.lastrowid}), 201


@admin_bp.route('/company-accounts/<int:user_id>', methods=['PUT'])
@admin_required
def update_company_account(user_id):
    data = request.get_json()
    db = get_db()
    user = db.execute("SELECT id FROM users WHERE id = ? AND role = 'company'", [user_id]).fetchone()
    if not user:
        return jsonify({'error': 'Company account not found'}), 404

    if 'status' in data:
        db.execute('UPDATE users SET account_status=? WHERE id=?', [data['status'], user_id])
    if 'company_id' in data:
        db.execute('UPDATE users SET company_id=? WHERE id=?', [data['company_id'], user_id])
    db.commit()
    return jsonify({'message': 'Account updated'}), 200


@admin_bp.route('/company-accounts/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_company_account(user_id):
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ? AND role = 'company'", [user_id])
    db.commit()
    return jsonify({'message': 'Account deleted'}), 200


# ── Training Data Management ────────────────────────────────────────────────

@admin_bp.route('/training-data/years', methods=['GET'])
@admin_required
def training_data_years():
    db = get_db()
    rows = db.execute("""
        SELECT graduation_year, COUNT(*) AS count,
               SUM(employed) AS employed_count
        FROM ml_training_rows
        WHERE is_active = 1
        GROUP BY graduation_year
        ORDER BY graduation_year DESC
    """).fetchall()
    total = db.execute("SELECT COUNT(*) AS n FROM ml_training_rows WHERE is_active = 1").fetchone()['n']
    return jsonify({
        'years': [{'year': r['graduation_year'], 'count': r['count'], 'employed': r['employed_count']} for r in rows],
        'total': total,
    }), 200


def _sync_alumni_to_training_data(db):
    """Remove alumni accounts for years that have no training data (keeps in sync with datasets)."""
    ep = ','.join('?' * len(TEST_ACCOUNT_EMAILS))
    years_with_data = set(
        r['graduation_year'] for r in
        db.execute("SELECT DISTINCT graduation_year FROM ml_training_rows WHERE is_active=1").fetchall()
    )
    base = (f"DELETE FROM users WHERE role='alumni' "
            f"AND (is_test_account=0 OR is_test_account IS NULL) "
            f"AND LOWER(email) NOT IN ({ep})")
    if not years_with_data:
        removed = db.execute(base, list(TEST_ACCOUNT_EMAILS)).rowcount
    else:
        yp = ','.join('?' * len(years_with_data))
        removed = db.execute(
            base + f" AND graduation_year NOT IN ({yp})",
            list(TEST_ACCOUNT_EMAILS) + list(years_with_data)
        ).rowcount
    db.commit()
    return removed


@admin_bp.route('/training-data/by-year/<int:year>', methods=['DELETE'])
@admin_required
def delete_training_data_by_year(year):
    db = get_db()
    cur = db.execute("DELETE FROM ml_training_rows WHERE graduation_year = ?", [year])
    db.execute("DELETE FROM employment_data WHERE year = ?", [year])
    db.execute("DELETE FROM program_rates WHERE year = ?", [year])
    # Also delete alumni accounts tied to this graduation year
    alumni_cur = db.execute(
        "DELETE FROM users WHERE role='alumni' AND graduation_year=? AND (is_test_account=0 OR is_test_account IS NULL)",
        [year]
    )
    db.commit()
    return jsonify({
        'message': f'Deleted {cur.rowcount} training rows and {alumni_cur.rowcount} alumni accounts for {year}.',
        'deleted': cur.rowcount,
        'alumni_deleted': alumni_cur.rowcount,
        'forecast_updated': True,
    }), 200



# ── Bulk Alumni Import ──────────────────────────────────────────────────────

@admin_bp.route('/users/bulk-import', methods=['POST'])
@admin_required
def bulk_import_users():
    import bcrypt as _bcrypt
    import pandas as pd
    import io

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400

    dataset_year_raw = request.form.get('dataset_year', '').strip()
    year_fallback = None
    try:
        year_fallback = int(dataset_year_raw) if dataset_year_raw else None
    except ValueError:
        pass

    skip_email = _is_truthy(request.form.get('skip_email'), default=False)

    filename = file.filename.lower()
    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file.read()))
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(file.read()))
        else:
            return jsonify({'error': 'Only CSV and Excel files are supported'}), 400
    except Exception as e:
        return jsonify({'error': f'Could not parse file: {e}'}), 400

    df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

    def _find_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    email_col     = _find_col(['email', 'email_address', 'e-mail'])
    if not email_col:
        return jsonify({'error': "No 'email' column found in the file"}), 400

    name_col      = _find_col(['name', 'full_name', 'fullname', 'student_name'])
    course_col    = _find_col(['program', 'course', 'degree'])
    year_col      = _find_col(['graduation_year', 'jr_grad', 'grad_year', 'year_graduated', 'year'])
    alumni_id_col = _find_col(['alumni_id', 'student_id', 'id_number'])
    employed_col  = _find_col(['employed', 'employment_status', 'employment', 'is_employed', 'status'])
    grade_col     = _find_col(['cgpa', 'avg_grade', 'gpa', 'general_average'])
    soft_col      = _find_col(['soft_skills', 'soft_skills_avg', 'soft_skill', 'softskills'])
    hard_col      = _find_col(['hard_skills', 'hard_skills_avg', 'hard_skill', 'hardskills'])

    db = get_db()
    existing_map = {
        r['email'].lower(): {'id': r['id'], 'role': r['role']}
        for r in db.execute("SELECT id, role, email FROM users").fetchall()
        if r['email']
    }

    created, updated, skipped, failed = [], [], [], []

    for _, row in df.iterrows():
        email = str(row.get(email_col, '') or '').strip().lower()
        if not email or '@' not in email:
            skipped.append({'email': email or '(blank)', 'reason': 'Invalid email'})
            continue

        def _safe(col, cast, default):
            try:
                v = row.get(col) if col else None
                if v is None or str(v).strip() in ('', 'nan', 'NaN', 'NaT'):
                    return default
                return int(float(v)) if cast is int else cast(v)
            except Exception:
                return default

        graduation_year = _safe(year_col, int, year_fallback)
        if not graduation_year and alumni_id_col:
            aid = str(row.get(alumni_id_col, '') or '').strip()
            m = re.search(r'\b(20\d{2})\b', aid)
            if m:
                graduation_year = int(m.group(1))
        if not graduation_year:
            skipped.append({'email': email, 'reason': 'No graduation year'})
            continue

        first_name, last_name = 'Alumni', ''
        if name_col and row.get(name_col):
            raw = str(row[name_col]).strip()
            if ',' in raw:
                parts = raw.split(',', 1)
                last_name = parts[0].strip().title()
                fn_parts = parts[1].strip().split()
                first_name = fn_parts[0].title() if fn_parts else 'Alumni'
            else:
                parts = raw.split()
                first_name = parts[0].title() if parts else 'Alumni'
                last_name = parts[-1].title() if len(parts) > 1 else ''

        course      = _normalize_course_value(row.get(course_col)) if course_col else ''
        employed    = (_parse_employment_flag(row.get(employed_col)) if employed_col else None) or 0
        avg_grade   = _safe(grade_col, float, 0.0)
        soft_skills = _safe(soft_col, float, 0.0)
        hard_skills = _safe(hard_col, float, 0.0)

        password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        pw_hash  = _bcrypt.hashpw(password.encode(), _bcrypt.gensalt(rounds=4)).decode()

        try:
            existing = existing_map.get(email)
            if existing:
                if existing['role'] != 'alumni':
                    skipped.append({'email': email, 'reason': 'Non-alumni account exists'})
                    continue
                db.execute("""
                    UPDATE users SET first_name=?, last_name=?, course=?,
                      graduation_year=?, avg_grade=?, soft_skills=?, hard_skills=?, employed=?
                    WHERE id=?
                """, [first_name, last_name, course, graduation_year,
                      avg_grade, soft_skills, hard_skills, employed, existing['id']])
                updated.append({'email': email, 'name': f"{first_name} {last_name}".strip()})
                continue

            db.execute("""
                INSERT INTO users
                  (first_name, last_name, email, password_hash, role,
                   course, graduation_year, avg_grade, soft_skills, hard_skills,
                   employed, account_status)
                VALUES (?,?,?,?,'alumni',?,?,?,?,?,?,'Active')
            """, [first_name, last_name, email, pw_hash,
                  course, graduation_year, avg_grade, soft_skills, hard_skills, employed])

            if skip_email:
                email_sent, email_err = False, 'skipped'
            else:
                email_sent, email_err = _send_welcome_email(email, f"{first_name} {last_name}".strip(), password)
            created.append({'email': email, 'name': f"{first_name} {last_name}".strip(),
                            'email_sent': email_sent, 'email_error': email_err})
        except Exception as e:
            failed.append({'email': email, 'reason': str(e)})

    db.commit()
    return jsonify({
        'message': f'Import complete: {len(created)} created, {len(updated)} updated, {len(skipped)} skipped',
        'created': created, 'updated': updated, 'skipped': skipped, 'failed': failed,
    }), 200
